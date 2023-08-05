import shutil
from fbprophet.plot import plot_plotly
import base64
import html 
import tempfile
import platform
import xlwt
import tzlocal
import glob
import numpy as np
from pandas.io.json import json_normalize
import pandas
import requests
import json
import urllib.request, urllib.parse, urllib.error
from collections import Counter
from urllib.error import HTTPError
from datetime import datetime, timedelta
import re
from easydict import EasyDict as edict
import time
import webbrowser
import os
from dateutil.parser import parse
import sys
import configparser
from openpyxl import Workbook
from openpyxl.styles import Alignment
from perfecto.perfectoactions import create_summary
from openpyxl.reader.excel import load_workbook

"""
    This is the payload based on only start and end date
"""
TEMP_DIR = "/tmp" if platform.system() == "Darwin" else tempfile.gettempdir()

class my_dictionary(dict):  
  
    # __init__ function  
    def __init__(self):  
        self = dict()  
          
    # Function to add key:value  
    def add(self, key, value):  
        self[key] = value  

def payloadJobAll(oldmilliSecs, current_time_millis, jobName, jobNumber, page, boolean):
    payload = my_dictionary()
    if oldmilliSecs != 0 : payload.add("startExecutionTime[0]", oldmilliSecs) 
    if current_time_millis != 0 : payload.add("endExecutionTime[0]", current_time_millis)
    payload.add("_page", page)
    if jobName != "":
        for i, job in enumerate(jobName.split(";")):
            payload.add("jobName[" +  str(i) + "]", job)
    if jobNumber != "" and boolean:
        for i, job in enumerate(jobName.split(",")):
                payload.add("jobNumber[" +  str(i) + "]", jobNumber)
    return payload

"""
    Retrieve a list of test executions within the last month
    :return: JSON object contains the executions
"""
def retrieve_tests_executions(daysOlder, page):
    current_time_millis = 0
    oldmilliSecs = 0
    if endDate != "" :
        endTime = datetime.strptime(str(endDate) + " 23:59:59,999", "%Y-%m-%d %H:%M:%S,%f")
        print("endExecutionTime: " + str(endTime))
        millisec = endTime.timestamp() * 1000
        current_time_millis = round(int(millisec))
    if startDate != "":
        oldmilliSecs = pastDateToMS(startDate, daysOlder)
    if jobNumber != "" and jobName != "" and startDate != "" and endDate != "" :
        payload = payloadJobAll(oldmilliSecs, current_time_millis, jobName, jobNumber, page, False)
    else:
        payload = payloadJobAll(oldmilliSecs, current_time_millis, jobName, jobNumber, page, True)
    # creates http geat request with the url, given parameters (payload) and header (for authentication)
    r = requests.get(
        api_url, params=payload, headers={"PERFECTO_AUTHORIZATION": OFFLINE_TOKEN}
    )
    # print entire response
    # #print(str(r.content))
    print(str(r.url))
    return r.content


"""
    sends API request
"""


def send_request(url):
    try:
        response = urllib.request.urlopen(url)
    except HTTPError as e:
        content = e.read()
        response = content
    return response


"""
    sends API request and gets a json response
"""


def send_request_with_json_response(url):
    response = send_request(url)
    text = response.read().decode("utf-8")
    map = json.loads(text)
    return map


def flatten_json(nested_json, exclude=[""]):
    """Flatten json object with nested keys into a single level.
        Args:
            nested_json: A nested json object.
            exclude: Keys to exclude from output.
        Returns:
            The flattened json object if successful, None otherwise.
    """
    out = {}

    def flatten(x, name="", exclude=exclude):
        if type(x) is dict:
            for a in x:
                if a not in exclude:
                    flatten(x[a], name + a + "/")
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + "/")
                i += 1
        else:
            out[name[:-1]] = x

    flatten(nested_json)
    return out


def get_final_df(files):
    df = pandas.DataFrame()
    for file in files:
        if "csv" in xlformat:
            df = df.append(pandas.read_csv(file, low_memory=False))
        else:
            df = df.append(pandas.read_excel(file))
    return df

"""
   gets the top failed device pass count, handset errors and device/ desktop details
"""

def getDeviceDetails(device, deviceFailCount):
    devicePassCount = 0
    errorsCount = 0
    i = 0


    
    for resource in resources:
        try:
            test_execution = resource  # retrieve a test execution
            # get devices which fails
            platforms = test_execution["platforms"]  # retrieve the platforms
            platform = platforms[0]
            actual_deviceID = platform["deviceId"]
            if actual_deviceID in device:
                status = test_execution["status"]
                if status in "PASSED":
                    devicePassCount += 1
                elif status in "FAILED":
                    message = test_execution["message"]
                    if "HANDSET_ERROR" in message:
                        errorsCount += 1
                deviceType = platform["deviceType"]
                if "DESKTOP" in deviceType:
                    browserInfo = platform["browserInfo"]
                    topDeviceFailureDict[device] = [
                        platform["os"] + "_" + platform["osVersion"],
                        browserInfo["browserType"]
                        + "_"
                        + browserInfo["browserVersion"],
                        devicePassCount,
                        deviceFailCount,
                        errorsCount,
                    ]
                else:
                    mobileInfo = platform["mobileInfo"]
                    topDeviceFailureDict[device] = [
                        mobileInfo["manufacturer"],
                        mobileInfo["model"],
                        devicePassCount,
                        deviceFailCount,
                        errorsCount,
                    ]
        except IndexError:
            continue
        except KeyError:
            continue

        printProgressBar(
            i + 1,
            len(resources),
            prefix="Fetching Device details in Progress:",
            suffix="Complete",
            length=50,
        )
        i += 1


"""
   gets the total pass count of each failed case
"""


def getPassCount(testName):
    testNamePassCount = 0
    i = 0
    for resource in resources:
        try:
            test_execution = resource  # retrieve a test execution
            name = test_execution["name"]

            if testName in name:
                status = test_execution["status"]
                if status in "PASSED":
                    testNamePassCount += 1
        except IndexError:
            continue
        except KeyError:
            continue
        printProgressBar(
            i + 1,
            len(resources),
            prefix="Pass % calculation API in Progress:",
            suffix="Complete",
            length=50,
        )
        i += 1
    return testNamePassCount


"""
   gets fail and pass count of each test case and assigns it to a dict
"""


def getTCDetails(tcName, failureCount):
    topTCFailureDict[tcName] = [failureCount, getPassCount(tcName)]


"""
   calculates the percetage of a part and whole number
"""


def percentageCalculator(part, whole):
    if int(whole) > 0:
        calc = (100 * float(part) / float(whole), 0)
        calc = round(float((calc[0])), 2)
    else:
        calc = 0
    return calc


"""
   gets start date to milliseconds
"""


def pastDateToMS(startDate, daysOlder):
    dt_obj = datetime.strptime(
        startDate + " 00:00:00,00", "%Y-%m-%d %H:%M:%S,%f"
    ) - timedelta(days=daysOlder)
    print("startExecutionTime: " + str(dt_obj))
    millisec = dt_obj.timestamp() * 1000
    oldmilliSecs = round(int(millisec))
    return oldmilliSecs


"""
   gets past pass percentage of tests
"""


def pastPassPercentageCalculator(daysOlder):
    totalPassCount = 0
    totalTCCount = 0
    i = 0
    for resource in resources:
        try:
            totalTCCount += 1
            test_execution = resource  # retrieve a test execution
            status = test_execution["status"]
            if status in "PASSED":
                totalPassCount += 1
        except IndexError:
            continue
        except KeyError:
            continue
        printProgressBar(
            i + 1,
            len(resources),
            prefix="TC Pass % based on older days API in Progress:",
            suffix="Complete",
            length=50,
        )
        i += 1
    return str(percentageCalculator(totalPassCount, totalTCCount))


def color_negative_red(value):
    # if "PASSED" in value:
    #     color = 'green'
    # elif "FAILED" in value:
    #     color = 'red'
    # else:
    #     color = 'grey'
    color = 'red' if value < 1 else 'black'
    return 'color: %s' % color



"""
   gets' Perfecto reporting API responses, creates dict for top device failures, auto suggestions and top tests failures and prepared json
"""

def prepareReport(jobName, jobNumber):
    page = 1
    i = 0
    truncated = True
    failureList = {}
    cleanedFailureList = {}
    device_Dictionary = {}
    totalFailCount = 0
    totalPassCount = 0
    totalUnknownCount = 0
    totalTCCount = 0
    labIssuesCount = 0
    scriptingIssuesCount = 0
    orchestrationIssuesCount = 0
    testNameFailureList = {}
    suggesstionsDict = {}
    global topDeviceFailureDict
    global topTCFailureDict
    global resources
    failureList.clear()
    cleanedFailureList.clear()
    testNameFailureList.clear()
    device_Dictionary.clear()
    resources.clear()
    suggesstionsDict.clear()
    while truncated == True:
        print(
            "Retrieving all the test executions in your lab. Current page: "
            + str(page)
            + ". Hold On!!"
        )
        executions = retrieve_tests_executions(0, page)
        # print(executions)
        # Loads JSON string into JSON object
        executions = json.loads(executions)
        if "{'userMessage': 'Failed decoding the offline token:" in str(executions):
            raise Exception("please change the offline token for your cloud")
        if "userMessage': 'Missing Perfecto-TenantId header" in str(executions):
            raise Exception("Check the cloud name and security tokens")
        try:
            executionList = executions["resources"]
        except TypeError as e:
            print(executions)
            raise Exception("Unable to find matching records for: " + str(criteria) + ", error:" + str(executions['userMessage']))
            sys.exit(-1)
        if len(executionList) == 0:
            print("0 test executions")
            break
        else:
            # print(str(executions))
            metadata = executions["metadata"]
            truncated = metadata["truncated"]
            if page >= 1:
                resources.extend(executionList)
            else:
                resources.append(executionList)
            page += 1
    if len(resources) > 0:
        jsonDump = json.dumps(resources)
        resources = json.loads(jsonDump)
        totalTCCount = len(resources)
        print("Total executions: " + str(len(resources)))
        df = pandas.DataFrame([flatten_json(x) for x in resources])
        df["startTime"] = pandas.to_datetime(df["startTime"].astype(int), unit="ms")
        df["startTime"] = (
            df["startTime"].dt.tz_localize("utc").dt.tz_convert(tzlocal.get_localzone())
        )
        df["startTime"] = df["startTime"].dt.strftime("%d/%m/%Y %H:%M:%S")
        df.loc[df['endTime'] < 1, 'endTime'] = int(round(time.time() * 1000)) 
        df["endTime"] = pandas.to_datetime(df["endTime"].astype(int), unit="ms")
        df["endTime"] = (
            df["endTime"].dt.tz_localize("utc").dt.tz_convert(tzlocal.get_localzone())
        )
        df["endTime"] = df["endTime"].dt.strftime("%d/%m/%Y %H:%M:%S")
        if "month" not in df.columns:
            df["month"] = pandas.to_datetime(df["startTime"], format='%d/%m/%Y %H:%M:%S').dt.to_period('M')
        if "startDate" not in df.columns:
            df['startDate'] = pandas.to_datetime(pandas.to_datetime(df["startTime"], format='%d/%m/%Y %H:%M:%S').dt.to_period('D').astype(str))
        if "week" not in df.columns:
            df['week'] = pandas.to_datetime(df['startDate'].dt.strftime("%Y/%m/%d")) - df['startDate'].dt.weekday.astype('timedelta64[D]')
        if "Duration" not in df.columns:
            df["Duration"] = pandas.to_datetime(df["endTime"]) - pandas.to_datetime(
                df["startTime"]
            )
            df["Duration"] = df["Duration"].dt.seconds
            df["Duration"] = pandas.to_datetime(df["Duration"], unit='s').dt.strftime("%H:%M:%S")
        if "failureReasonName" not in df.columns: 
            df["failureReasonName"] = ""
        # df["name"] = '=HYPERLINK("'+df["reportURL"]+'", "'+df["name"]+'")'  # has the ability to hyperlink name in csv'

        #Filter only job and job number if dates are parameterized as well but show full histogram
        if jobNumber != "" and jobName != "":
            ori_df = df
            df = df[df['job/number'].astype(str) == jobNumber]
        if startDate != "":
            name = startDate
        else:
            name = jobName + '_' + jobNumber
        df_to_xl(df, str(name).replace("/","_"))
    os.chdir(".")
    files = glob.glob('*.{}'.format(xlformat))
    if consolidate != "":
        for file in files:
            if os.path.isfile(file):
                shutil.copy2(file, consolidate)
        files = glob.iglob(os.path.join(consolidate, "*." + xlformat))
    df = get_final_df(files)
    df = df.sort_values(by=['startDate'], ascending=False)
    if jobNumber != "" and jobName != "":
        ori_df = df
        df = df[df['job/name'].astype(str).isin(jobName.split(";"))]
        df = df[df['job/number'].round(0).astype(int).isin(jobNumber.split(";"))]
    if jobNumber == "" and jobName != "":
        ori_df = df
        df = df[df['job/name'].astype(str).isin(jobName.split(";"))]
    df_to_xl(df, "final")   
    if (len(df)) < 1:
        print("Unable to find any test executions for the criteria: " + criteria)
        sys.exit(-1)
    import plotly.express as px
    import plotly
    #ggplot2 #plotly_dark #simple_white
    graphs = []
    graphs.append('<div id="nestle-section">')
    counter = 8
    with open(live_report_filename, 'a') as f:
        f.write('<div id="nestle-section">')
    duration = "weeks"
    if startDate != "":
        delta = datetime.strptime(endDate, "%Y-%m-%d") - datetime.strptime(startDate, "%Y-%m-%d")
        if (delta.days) <= 14:
            duration = "dates"
    else:
        duration = "dates"
    joblist = []
    if "job/name" in df.columns and jobName != "":
        joblist = sorted(df['job/name'].dropna().unique())
    else:
        joblist.append("Overall!")
    for job in joblist: 
        predict_df = df
        fig = []
        if job != "Overall!":
            if job in jobName:
                if duration == "dates":
                    fig = px.histogram(df.loc[df['job/name'] == job], x="startDate", color="status", color_discrete_map= {"PASSED":"limegreen","FAILED":"crimson","UNKNOWN":"#9da7f2","BLOCKED":"#e79a00"}, hover_data=df.columns, template="seaborn", opacity=0.5)  
                else:
                    fig = px.histogram(df.loc[df['job/name'] == job], x="week", color="status",
                                    hover_data=df.columns, color_discrete_map= {"PASSED":"limegreen","FAILED":"crimson","UNKNOWN":"#9da7f2","BLOCKED":"#e79a00"}, template="seaborn", opacity=0.5)
                predict_df = df.loc[df['job/name'] == job]
        else:
            fig = px.histogram(df, x="startDate", color="status", color_discrete_map= {"PASSED":"limegreen","FAILED":"crimson","UNKNOWN":"#9da7f2","BLOCKED":"#e79a00"}, hover_data=df.columns, template="seaborn", opacity=0.5)    
        predict_df = predict_df.groupby(['startDate']).size().reset_index(name='#status').sort_values('#status', ascending=False)
        if fig:   
            fig = update_fig(fig, "histogram", job, duration)
            encoded = base64.b64encode(plotly.io.to_image(fig))
            graphs.append('<input type="radio" id="tab' + str(counter) + '" name="tabs" checked=""/><label for="tab' + str(counter) + '">Trends: ' + job + '</label><div class="tab-content1"><img src="data:image/png;base64, {}"'.format(encoded.decode("ascii")) + " alt='days or weeks summary of " + job + "' id='reportDiv' onClick='zoom(this)'></img></div>")
            with open(live_report_filename, 'a') as f:
                f.write('<input type="radio" id="tab' + str(counter) + '" name="tabs" checked=""/><label for="tab' + str(counter) + '">Trends: ' + job + ' </label><div class="tab-content1">' + fig.to_html(full_html=False, include_plotlyjs='cdn') + '</div>')
        if job == "Overall!" or job in jobName:
            if len(predict_df.index) > 1:
                predict_df = predict_df.rename(columns={'startDate': 'ds', '#status' : 'y'})
                predict_df['cap'] = (int(predict_df['y'].max()) * 2)
                predict_df['floor'] = 0
                from fbprophet import Prophet
                with suppress_stdout_stderr():
                    m = Prophet(seasonality_mode='additive', growth='logistic', changepoint_prior_scale = 0.001 ).fit(predict_df, algorithm='Newton')
                future = m.make_future_dataframe(periods=30)
                future['cap'] = (int(predict_df['y'].max()) * 2)
                future['floor'] = 0
                forecast = m.predict(future)
                forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail()
                fig = plot_plotly(m, forecast)
                fig = update_fig(fig, "prediction", job, duration)
                encoded = base64.b64encode(plotly.io.to_image(fig))
                counter+=1
                graphs.append('<input type="radio" id="tab' + str(counter) + '" name="tabs" checked=""/><label for="tab' + str(counter) + '">Monthly Prediction: ' + job + '</label><div class="tab-content1"><div class="reportDiv"><img src="data:image/png;base64, {}"'.format(encoded.decode("ascii")) + " alt='prediction summary' id='reportDiv' onClick='zoom(this)'></img></div></p></div>")
                with open(live_report_filename, 'a') as f:
                    f.write('<input type="radio" id="tab' + str(counter) + '" name="tabs" checked=""/><label for="tab' + str(counter) + '">Monthly Prediction: ' + job + '</label><div class="tab-content1"><div class="predictionDiv">' + fig.to_html(full_html=False, include_plotlyjs='cdn') + " </img></div></p></div>")
            else:
                print("Note: AI Prediction for job: " + job + " requires more than 2 days of data to analyze!")
        counter+=1
    graphs.append('</div>')
    with open(live_report_filename, 'a') as f:
        f.write('</div>')
    for resource in resources:
        try:
            #            totalTCCount += 1
            test_execution = resource  # retrieve a test execution
            status = test_execution["status"]
            if status in "FAILED":
                totalFailCount += 1
                # get failed test names
                name = test_execution["name"]
                if name in testNameFailureList:
                    testNameFailureList[name] += 1
                else:
                    testNameFailureList[name] = 1
                # get devices which fails
                platforms = test_execution["platforms"]  # retrieve the platforms
                platform = platforms[0]
                actual_deviceID = platform["deviceId"]
                if actual_deviceID in device_Dictionary:
                    device_Dictionary[actual_deviceID] += 1
                else:
                    device_Dictionary[actual_deviceID] = 1
                # get error messages
                message = test_execution["message"]
                if message in failureList:
                    failureList[message] += 1
                else:
                    failureList[message] = 1
            elif status in "PASSED":
                totalPassCount += 1
            elif status in "UNKNOWN":
                totalUnknownCount += 1
        except IndexError:
            continue
        except KeyError:
            continue

    # Top 5 device which failed
    # topDeviceFailureDict.clear()
    # deviceFailureDict = Counter(device_Dictionary)
    # for device, deviceFailCount in deviceFailureDict.most_common(5):
    #     getDeviceDetails(device, deviceFailCount)

    # Top 5 failure tests along with pass count
    topTCFailureDict.clear()
    testNameFailureListDict = Counter(testNameFailureList)
    for failedTestName, failedTestNamesCount in testNameFailureListDict.most_common(5):
        getTCDetails(str(failedTestName), failedTestNamesCount)

    # last 7/14 days pass % calculator
    # print("14D" + str(pastPassPercentageCalculator(14)))
    print("total unique fail count:" + str(len(failureList)))
    i = 0
    failureListFileName = CQL_NAME + "_failures" + ".txt"
    print(
        "transfering all failure reasons to: %s"
        % (os.path.join(os.path.abspath(os.curdir), failureListFileName))
    )
    open(failureListFileName, "a").close
    for commonError, commonErrorCount in failureList.items():
        for labIssue in labIssues:
            if re.search(labIssue, commonError):
                labIssuesCount += commonErrorCount
                break
        for orchestrationIssue in orchestrationIssues:
            if re.search(orchestrationIssue, commonError):
                orchestrationIssuesCount += commonErrorCount
                break
        error = commonError
        regEx_Filter = "Build info:|For documentation on this error|at org.xframium.page|Scenario Steps:| at WebDriverError|\(Session info:|XCTestOutputBarrier\d+|\s\tat [A-Za-z]+.[A-Za-z]+.|View Hierarchy:|Got: |Stack Trace:|Report Link|at dalvik.system|Output:\nUsage|t.*Requesting snapshot of accessibility"
        if re.search(regEx_Filter, error):
            error = str(re.compile(regEx_Filter).split(error)[0])
            if "An error occurred. Stack Trace:" in error:
                error = error.split("An error occurred. Stack Trace:")[1]
        if re.search("error: \-\[|Fatal error:", error):
            error = str(re.compile("error: \-\[|Fatal error:").split(error)[1])
        if error.strip() in cleanedFailureList:
            cleanedFailureList[error.strip()] += 1
        else:
            cleanedFailureList[error.strip()] = commonErrorCount
        scriptingIssuesCount = totalFailCount - (
            orchestrationIssuesCount + labIssuesCount
        )
        with open(failureListFileName, "a", encoding="utf-8") as myfile:
            myfile.write(
                error.strip() + "\n*******************************************\n"
            )
        printProgressBar(
            i + 1,
            len(failureList),
            prefix="chart preparation in Progress:",
            suffix="Complete",
            length=50,
        )
        i += 1

    # Top 5 failure reasons
    topFailureDict = {}

    failureDict = Counter(cleanedFailureList)
    for commonError, commonErrorCount in failureDict.most_common(5):
        topFailureDict[commonError] = int(commonErrorCount)

    # reach top errors and clean them
    i = 0
    for commonError, commonErrorCount in topFailureDict.items():
        if "Device not found" in error:
            error = (
                "Raise a support case as *|*"
                + commonError.strip()
                + "*|* as it occurs in *|*"
                + str(commonErrorCount)
                + "*|* occurrences"
            )
        elif "Cannot open device" in error:
            error = (
                "Reserve the device/ use perfecto lab auto selection feature to avoid:  *|*"
                + commonError.strip()
                + "*|* as it occurs in *|*"
                + str(commonErrorCount)
                + "*|* occurrences"
            )
        else:
            error = (
                "Fix the error: *|*"
                + commonError.strip()
                + "*|* as it occurs in *|*"
                + str(commonErrorCount)
                + "*|* occurrences"
            )
        suggesstionsDict[error] = commonErrorCount
        printProgressBar(
            i + 1,
            len(topFailureDict),
            prefix="Generation of suggestions in Progress:",
            suffix="Complete",
            length=50,
        )
        i += 1
    eDict = edict(
        {
            "last24h": int(percentageCalculator(totalPassCount, totalTCCount)),
            "lab": labIssuesCount,
            "orchestration": orchestrationIssuesCount,
            "scripting": scriptingIssuesCount,
            "unknowns": totalUnknownCount,
            "executions": totalTCCount,
            "recommendations": [
                {
                    "rank": 1,
                    "recommendation": "-",
                    "impact": 0,
                    "impactMessage": "null",
                },
                {
                    "rank": 2,
                    "recommendation": "-",
                    "impact": 0,
                    "impactMessage": "null",
                },
                {
                    "rank": 3,
                    "recommendation": "-",
                    "impact": 0,
                    "impactMessage": "null",
                },
                {
                    "rank": 4,
                    "recommendation": "-",
                    "impact": 0,
                    "impactMessage": "null",
                },
                {
                    "rank": 5,
                    "recommendation": "-",
                    "impact": 0,
                    "impactMessage": "null",
                },
            ],
            "topProblematicDevices": [
                {
                    "rank": 1,
                    "model": "",
                    "os": "",
                    "id": "",
                    "passed": 0,
                    "failed": 0,
                    "errors": 0,
                },
                {
                    "rank": 2,
                    "model": "",
                    "os": "",
                    "id": "",
                    "passed": 0,
                    "failed": 0,
                    "errors": 0,
                },
                {
                    "rank": 3,
                    "model": "",
                    "os": "",
                    "id": "",
                    "passed": 0,
                    "failed": 0,
                    "errors": 0,
                },
                {
                    "rank": 4,
                    "model": "",
                    "os": "",
                    "id": "",
                    "passed": 0,
                    "failed": 0,
                    "errors": 0,
                },
                {
                    "rank": 5,
                    "model": "",
                    "os": "",
                    "id": "",
                    "passed": 0,
                    "failed": 0,
                    "errors": 0,
                },
            ],
            "topFailingTests": [
                {"rank": 1, "test": "", "failures": 0, "passes": 0},
                {"rank": 2, "test": "", "failures": 0, "passes": 0},
                {"rank": 3, "test": "", "failures": 0, "passes": 0},
                {"rank": 4, "test": "", "failures": 0, "passes": 0},
                {"rank": 5, "test": "", "failures": 0, "passes": 0},
            ],
        }
    )
    jsonObj = edict(eDict)

    if float(percentageCalculator(totalUnknownCount, totalTCCount)) >= 30:
        suggesstionsDict[
            "# Fix the unknowns. The unknown script ratio is too high (%) : "
            + str(percentageCalculator(totalUnknownCount, totalTCCount))
            + "%"
        ] = percentageCalculator(
            totalPassCount + totalUnknownCount, totalTCCount
        ) - percentageCalculator(
            totalPassCount, totalTCCount
        )
    if len(suggesstionsDict) < 5:
        if (len(topTCFailureDict)) > 1:
            for tcName, status in topTCFailureDict.items():
                suggesstionsDict[
                    "# Fix the top failing test: "
                    + tcName
                    + " as the failures count is: "
                    + str(int((str(status).split(",")[0]).replace("[", "").strip()))
                ] = 1
                break
    if len(suggesstionsDict) < 5:
        if (len(topDeviceFailureDict)) > 1:
            for device, status in topDeviceFailureDict.items():
                if "_" in str(status):
                    suggesstionsDict[
                        "# Fix the issues with top failing desktop: "
                        + (str(status).split(",")[0])
                        .replace("[", "")
                        .replace("'", "")
                        .strip()
                        + " "
                        + (str(status).split(",")[1]).replace("'", "").strip()
                        + " as the failures count is: "
                        + str(int((str(status).split(",")[3]).strip()))
                    ] = 1
                else:
                    suggesstionsDict[
                        "# Fix the top failing device: "
                        + device
                        + " as the failures count is: "
                        + str(int((str(status).split(",")[3]).strip()))
                    ] = 1
                break
    if len(suggesstionsDict) < 5:
        if int(percentageCalculator(totalFailCount, totalTCCount)) > 15:
            if totalTCCount > 0:
                suggesstionsDict[
                    "# Fix the failures. The total failures % is too high (%) : "
                    + str(percentageCalculator(totalFailCount, totalTCCount))
                    + "%"
                ] = totalFailCount
    if len(suggesstionsDict) < 5:
        if float(percentageCalculator(totalPassCount, totalTCCount)) < 80 and (
            totalTCCount > 0
        ):
            suggesstionsDict[
                "# Fix the failures. The total pass %  is too less (%) : "
                + str(int(percentageCalculator(totalPassCount, totalTCCount)))
                + "%"
            ] = (
                100
                - (
                    percentageCalculator(
                        totalPassCount + totalUnknownCount, totalTCCount
                    )
                    - percentageCalculator(totalPassCount, totalTCCount)
                )
            ) - int(
                percentageCalculator(totalPassCount, totalTCCount)
            )
    if len(suggesstionsDict) < 5:
        if totalTCCount == 0:
            suggesstionsDict[
                "# There are no executions for today. Try Continuous Integration with any tools like Jenkins and schedule your jobs today. Please reach out to Professional Services team of Perfecto for any assistance :) !"
            ] = 100
        elif int(percentageCalculator(totalPassCount, totalTCCount)) > 80:
            suggesstionsDict["# Overall Pass% is " + str(int(percentageCalculator(totalPassCount, totalTCCount)))+ "! Keep it up!"] = 0

        int(percentageCalculator(totalFailCount, totalTCCount)) > 15
    print("**************#Top 5 failure reasons ")
    topSuggesstionsDict = Counter(suggesstionsDict)
    counter = 0
    for sugg, commonErrorCount in topSuggesstionsDict.most_common(5):
        impact = 1
        if sugg.startswith("# "):
            sugg = sugg.replace("# ", "")
            impact = commonErrorCount
        else:
            impact = percentageCalculator(
                totalPassCount + commonErrorCount, totalTCCount
            ) - percentageCalculator(totalPassCount, totalTCCount)
        jsonObj.recommendations[counter].impact = int(impact)
        if int(impact) < 1:
            jsonObj.recommendations[counter].recommendation = (
                sugg.replace('"', "*|*").replace("'", "*|*").strip()
                + ". Impact: "
                + str(("%.2f" % round(impact, 2)))
                + "%"
            )
        else:
            jsonObj.recommendations[counter].recommendation = (
                sugg.replace('"', "*|*").replace("'", "*|*").strip()
            )
        print(str(counter + 1) + "." + str(sugg))
        printProgressBar(
            counter + 1,
            5,
            prefix="Top suggesstions in Progress:",
            suffix="Complete",
            length=5,
        )
        counter += 1
    counter = 0
    print("**************#Top 5 device which failed")
    for device, status in topDeviceFailureDict.items():
        print(str(counter + 1) + "." + device, status)
        jsonObj.topProblematicDevices[counter].id = device.strip()
        jsonObj.topProblematicDevices[counter].os = (
            (str(status).split(",")[0]).replace("[", "").replace("'", "").strip()
        )
        jsonObj.topProblematicDevices[counter].model = (
            (str(status).split(",")[1]).replace("'", "").strip()
        )
        jsonObj.topProblematicDevices[counter].passed = int(
            (str(status).split(",")[2]).strip()
        )
        jsonObj.topProblematicDevices[counter].failed = int(
            (str(status).split(",")[3]).strip()
        )
        jsonObj.topProblematicDevices[counter].errors = int(
            (str(status).split(",")[4]).replace("]", "").strip()
        )
        printProgressBar(
            counter + 1,
            5,
            prefix="Top device suggesstions in Progress:",
            suffix="Complete",
            length=5,
        )
        counter += 1
    df_model = pandas.DataFrame(jsonObj.topProblematicDevices)
    df_model["model"].replace("", np.nan, inplace=True)
    df_model.dropna(subset=["model"], inplace=True)
    counter = 0
    print("**************#Top 5 failure tests along with pass count")
    for tcName, status in topTCFailureDict.items():
        print(str(counter + 1) + "." + tcName, status)
        jsonObj.topFailingTests[counter].test = tcName.strip()
        jsonObj.topFailingTests[counter].failures = int(
            (str(status).split(",")[0]).replace("[", "").strip()
        )
        jsonObj.topFailingTests[counter].passes = int(
            (str(status).split(",")[1]).replace("]", "").strip()
        )
        printProgressBar(
            counter + 1,
            5,
            prefix="Top TC failures in Progress:",
            suffix="Complete",
            length=5,
        )
        counter += 1
    df2 = pandas.DataFrame(jsonObj.topFailingTests)
    df2["test"].replace("", np.nan, inplace=True)
    df2.dropna(subset=["test"], inplace=True)
    jsonObj = (
        str(jsonObj).replace("'", '"').replace('"null"', "null").replace("*|*", "'")
    )
    return graphs, df


"""
   shows the progress bar
"""

def as_text(value):
    """as texts"""
    if value is None:
        return ""
    return str(value)

def printProgressBar(
    iteration, total, prefix="", suffix="", decimals=1, length=10, fill="#"
):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
    """
    # percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    # filledLength = int(length * iteration // total)
    # bar = fill * filledLength + '-' * (length - filledLength)
    # print('\r%s |%s| %s%% %s' % (prefix, bar, percent, suffix), end = '\r')
    # Print New Line on Complete


#    if iteration == total:
#        print()

class suppress_stdout_stderr(object):
    '''
    A context manager for doing a "deep suppression" of stdout and stderr in
    Python, i.e. will suppress all print, even if the print originates in a
    compiled C/Fortran sub-function.
       This will not suppress raised exceptions, since exceptions are printed
    to stderr just before a script exits, and after the context manager has
    exited (at least, I think that is why it lets exceptions through).

    '''
    def __init__(self):
        # Open a pair of null files
        self.null_fds = [os.open(os.devnull, os.O_RDWR) for x in range(2)]
        # Save the actual stdout (1) and stderr (2) file descriptors.
        self.save_fds = (os.dup(1), os.dup(2))

    def __enter__(self):
        # Assign the null pointers to stdout and stderr.
        os.dup2(self.null_fds[0], 1)
        os.dup2(self.null_fds[1], 2)

    def __exit__(self, *_):
        # Re-assign the real stdout/stderr back to (1) and (2)
        os.dup2(self.save_fds[0], 1)
        os.dup2(self.save_fds[1], 2)
        # Close the null files
        os.close(self.null_fds[0])
        os.close(self.null_fds[1])

"""
   returns a boolean if the provided string is a date or nots
"""


def is_date(string):
    try:
        parse(string)
        return True
    except ValueError:
        return False

def df_to_xl(df, filename):
    custom_columns = [
        "name",
        "status",
        "platforms/0/os",
        "platforms/0/mobileInfo/model",
        "platforms/0/browserInfo/browserType",
        "platforms/0/browserInfo/browserVersion",
        "platforms/0/osVersion",
        "failureReasonName",
        "message",
        "startTime",
        "endTime",
        "Duration",
        "job/name",
        "job/number",
        "job/branch",
        "owner",
        "reportURL",
        "platforms/0/deviceId",
        "platforms/0/deviceType",
        "platforms/0/mobileInfo/manufacturer",
        "platforms/0/screenResolution",
        "platforms/0/location",
        "platforms/0/mobileInfo/imei",
        "platforms/0/mobileInfo/phoneNumber",
        "platforms/0/mobileInfo/distributor",
        "platforms/0/mobileInfo/firmware",
        "platforms/0/selectionCriteriaV2/0/name",
        "platforms/0/selectionCriteriaV2/1/name",
        "platforms/0/selectionCriteriaV2/2/name",
        "platforms/0/selectionCriteriaV2/2/value",
        "platforms/0/customFields/0/name",
        "platforms/0/customFields/0/value",
        "tags/0",
        "tags/1",
        "tags/2",
        "tags/3",
        "tags/4",
        "tags/5",
        "tags/6",
        "tags/7",
        "tags/8",
        "tags/9",        
        "tags/10",
        "tags/11",
        "tags/12",
        "tags/13",
        "tags/14",
        "tags/15",
        "tags/16",
        "id",
        "externalId",
        "uxDuration",
        "videos/0/startTime",
        "videos/0/endTime",
        "videos/0/format",
        "videos/0/streamingUrl",
        "videos/0/downloadUrl",
        "videos/0/screen/width",
        "videos/0/screen/height",
        "executionEngine/version",
        "project/name",
        "project/version",
        "automationFramework",
        "parameters/0/name",
        "parameters/0/value",
        "parameters/1/name",
        "parameters/1/value",
        "parameters/2/name",
        "parameters/2/value",
        "parameters/3/name",
        "parameters/3/value",
        "parameters/4/name",
        "parameters/4/value",
        "parameters/5/name",
        "parameters/5/value",
        "parameters/6/name",
        "parameters/6/value",
        "parameters/7/name",
        "parameters/7/value",
        "parameters/8/name",
        "parameters/8/value",
        "parameters/9/name",
        "parameters/9/value",
        "parameters/10/name",
        "parameters/10/value",
        "parameters/11/name",
        "parameters/11/value",
        "parameters/12/name",
        "parameters/12/value",        
        "parameters/13/name",
        "parameters/13/value",
        "platforms/0/mobileInfo/operator",
        "platforms/0/mobileInfo/operatorCountry",
        "platforms/0/selectionCriteriaV2/3/name",
        "platforms/0/selectionCriteriaV2/3/value",
        "platforms/0/selectionCriteriaV2/4/name",
        "platforms/0/selectionCriteriaV2/4/value",
        "platforms/0/selectionCriteriaV2/5/name",
        "platforms/0/selectionCriteriaV2/5/value",
        "platforms/0/selectionCriteriaV2/6/name",
        "platforms/0/selectionCriteriaV2/6/value",
        "platforms/0/selectionCriteriaV2/7/name",
        "platforms/0/selectionCriteriaV2/7/value",
        "customFields/0/name",
        "customFields/0/value",
        "customFields/1/name",
        "customFields/1/value",
        "artifacts/0/type",
        "artifacts/0/path",
        "artifacts/0/zipped",
        "artifacts/1/type",
        "artifacts/1/path",
        "artifacts/1/contentType",
        "artifacts/1/zipped",
        "artifacts/2/type",
        "artifacts/2/path",
        "artifacts/2/zipped",
        "artifacts/0/contentType",
        "artifacts/2/contentType",
        "platforms/1/deviceId",
        "platforms/1/deviceType",
        "platforms/1/os",
        "platforms/1/osVersion",
        "platforms/1/screenResolution",
        "platforms/1/location",
        "platforms/1/mobileInfo/imei",
        "platforms/1/mobileInfo/manufacturer",
        "platforms/1/mobileInfo/model",
        "platforms/1/mobileInfo/distributor",
        "platforms/1/mobileInfo/firmware",
        "platforms/1/selectionCriteriaV2/0/name",
        "platforms/1/selectionCriteriaV2/0/value",
        "platforms/1/customFields/0/name",
        "platforms/1/customFields/0/value",
        "videos/1/startTime",
        "videos/1/endTime",
        "videos/1/format",
        "videos/1/streamingUrl",
        "videos/1/downloadUrl",
        "videos/1/screen/width",
        "videos/1/screen/height",
        "platforms/1/mobileInfo/phoneNumber",
        "month",
        "week",
        "startDate",
    ]
    df = df[df.columns.intersection(custom_columns)]
    df = df.reindex(columns=custom_columns)
    df = df.dropna(axis=1, how="all")
    filename = [filename,".",xlformat]
    if "csv" in xlformat:
        df.to_csv("".join(filename), index=False)
    else:
        df.to_excel("".join(filename), index=False)    
    if "csv"  not in xlformat:
        wb = Workbook()
        wb = load_workbook("".join(filename))
        ws = wb.worksheets[0]
        for column_cells in ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5
        newfilename = os.path.abspath("".join(filename))
        wb.save(newfilename)

def get_report_details(item, temp, name, criteria):
    if name + "=" in item:
        temp = str(item).split("=")[1]
        criteria += "; " + name + ": " + temp 
    return temp, criteria

def update_fig(fig, type, job, duration):
    fig.update_layout(
    title={
        'text': '',
        'y':0.97,
        'x':0.5,
        'xanchor': 'center',
        'yanchor': 'top'},
    xaxis_title = duration,
    yaxis_title = "Test Status",
    font=dict(
        family = "Trebuchet MS, Helvetica, sans-serif",
        size = 12,
        color = 'black',
    ),
    autosize =True,
    hovermode = "x unified",
    yaxis={ 'tickformat' : '.0f' },
    xaxis_tickformat = '%d/%b/%y',
    ) 
    fig.update_yaxes(automargin=True)
    if type == "prediction":
        fig.update_layout( title={'text': ''}, yaxis_title = "Total tests executed",)
    return fig

def get_html_string(graphs):
    return (
        """
    <html lang="en">
       <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
    		     <head><title> Cloud Status</title>
          <script src="https://ajax.googleapis.com/ajax/libs/jquery/3.4.1/jquery.min.js"></script>
            <script>
              $(document).ready(function(){{
                document.getElementById("tabbed-device").click();
            }});
            $(document).ready(function(){{
                // Add smooth scrolling to all links
                $("a").on('click', function(event) {{

                    // Make sure this.hash has a value before overriding default behavior
                    if (this.hash !== "") {{
                    // Prevent default anchor click behavior
                    event.preventDefault();

                    // Store hash
                    var hash = this.hash;

                    // Using jQuery's animate() method to add smooth page scroll
                    // The optional number (800) specifies the number of milliseconds it takes to scroll to the specified area
                    $('html, body').animate({{
                        scrollTop: $(hash).offset().top
                    }}, 800, function(){{
                
                        // Add hash (#) to URL when done scrolling (default click behavior)
                        window.location.hash = hash;
                    }});
                    }} // End if
                }});
            }});
            $(document).ready(function(){{
              $("#myInput").on("keyup", function() {{
                var value = $(this).val().toLowerCase();
                $("#devicetable tbody tr").filter(function() {{
                  $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
                }});
              }});
            }});
                      $(document).ready(function(){{
              $("#myInput2").on("keyup", function() {{
                var value = $(this).val().toLowerCase();
                $("#usertable tbody tr").filter(function() {{
                  $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
                }});
              }});
            }});
            $(document).ready(function(){{
              $("#myInput3").on("keyup", function() {{
                var value = $(this).val().toLowerCase();
                $("#repotable tbody tr").filter(function() {{
                  $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
                }});
              }});
            }});
            </script>
    		<script type="text/javascript">
    	           $(document).ready(function(){{
                   $("#slideshow > div:gt(0)").show();
    				$("tbody tr:contains('Disconnected')").css('background-color','#fcc');
    				$("tbody tr:contains('ERROR')").css('background-color','#fcc');
    				$("tbody tr:contains('Un-available')").css('background-color','#fcc');
    				$("tbody tr:contains('Busy')").css('background-color','#fcc');
                    var table = document.getElementById("devicetable");
    				var rowCount = table.rows.length;
    				for (var i = 0; i < rowCount; i++) {{
    					if ( i >=1){{
                        available_column_number = 0;
                        device_id_column_number = 1;
    						if (table.rows[i].cells[available_column_number].innerHTML == "Available") {{
                                for(j = 0; j < table.rows[0].cells.length; j++) {{
    								table.rows[i].cells[j].style.backgroundColor = '#e6fff0';
                                        if(j=table.rows[0].cells.length){{
                                                if (table.rows[i].cells[(table.rows[0].cells.length - 1)].innerHTML.indexOf("failed") > -1) {{
                                                        table.rows[i].cells[j].style.color = '#660001';
                                                        table.rows[i].cells[j].style.backgroundColor = '#FFC2B5';
                                                }}
    							}}
                                 }}
    							var txt = table.rows[i].cells[device_id_column_number].innerHTML;
    							var url = 'http';
    							var row = $('<tr></tr>')
    							var link = document.createElement("a");
    							link.href = url;
    							link.innerHTML = txt;
    							link.target = "_blank";
    							table.rows[i].cells[device_id_column_number].innerHTML = "";
    							table.rows[i].cells[device_id_column_number].appendChild(link);
    						}}else{{
    							for(j = 0; j < table.rows[0].cells.length; j++) {{
    								table.rows[i].cells[j].style.color = '#660001';
                                         table.rows[i].cells[j].style.backgroundColor = '#FFC2B5';
    							}}
    						}}
    					}}
    				}}
                 }});
                 function myFunction() {{
                  var x = document.getElementById("myTopnav");
                  if (x.className === "topnav") {{
                    x.className += " responsive";
                  }} else {{
                    x.className = "topnav";
                  }}
                }}
                function zoom(element) {{
				         var data = element.getAttribute("src");
						 let w = window.open('about:blank');
						 let image = new Image();
						 image.src = data;
						 setTimeout(function(){{
						   w.document.write(image.outerHTML);
						 }}, 0);
				     }}
                function autoselect(element) {{
                     var data = element.getAttribute("id");
                     document.getElementById(data + "-1").checked = true;
                }}     
    		</script>

    		<meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
            <style>

                html {{
                height:100%;
                }}
                
                .tabbed {{
                display:  flex;
                text-align: left;
                flex-wrap: wrap;
                box-shadow: 0 0 80px rgba(101, 242, 183, 0.4);
                font-size: 12px;
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                }}
                .tabbed > input {{
                display: none;
                }}
                .tabbed > input:checked + label {{
                font-size: 14px;
                text-align: center;
                color: white;
                background-image: linear-gradient(to left, #bfee90, #333333, black,  #333333, #bfee90);
                }}
                .tabbed > input:checked + label + div {{
                color:darkslateblue;
                display: block;
                }}
                .tabbed > label {{
                background-image: linear-gradient(to left, #fffeea,  #333333, #333333 ,#333333 ,#333333 , #333333, #fffeea);
                color: white;
                text-align: center;
                display: block;
                order: 1;
                flex-grow: 1;
                padding: .3%;
                }}
                .tabbed > div {{
                order: 2;
                flex-basis: 100%;
                display: none;
                padding: 10px;
                }}

                /* For presentation only */
                .container {{
                width: 100%;
                margin: 0 auto;
                background-color: black;
                box-shadow: 0 0 20px rgba(400, 99, 228, 0.4);
                }}

                .tabbed {{
                border: 1px solid;
                }}

                hr {{
                background-color: white;
                height: 5px;
                border: 0;
                margin: 10px 0 0;
                }}
                
                hr + * {{
                margin-top: 10px;
                }}
                
                hr + hr {{
                margin: 0 0;
                }}

                .mystyle {{
                    font-size: 12pt;
                    font-family: "Trebuchet MS", Helvetica, sans-serif;
                    border-collapse: collapse;
                    border: 2px solid black;
                    margin:auto;
                    box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
                    background-color: #fffffa;
                }}

                .mystyle body {{
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                    table-layout: auto;
                    position:relative;
                }}

                #slide{{
                width:100%;
                height:auto;
                }}

                #myInput, #myInput2, #myInput3 {{
                background-image: url('http://www.free-icons-download.net/images/mobile-search-icon-94430.png');
                background-position: 2px 4px;
                background-repeat: no-repeat;
                background-size: 25px 30px;
                width: 40%;
                height:auto;
                font-weight: bold;
                font-size: 12px;
                padding: 11px 20px 12px 40px;
                box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
                }}

                p {{
                text-align:center;
                color:white;
                }}

                body {{
                background-color: rgba(185, 240, 218, 0.4);
                height: 100%;
                background-repeat:  repeat-y;
                background-position: right;
                background-size:  contain;
                background-attachment: initial;
                opacity:.93;
                }}

                h4 {{
                font-family:monospace;
                }}

                @keyframes slide {{
                0% {{
                    transform:translateX(-25%);
                }}
                100% {{
                    transform:translateX(25%);
                }}
                }}

                .mystyle table {{
                    table-layout: auto;
                    width: 100%;
                    height: 100%;
                    position:relative;
                    border-collapse: collapse;
                }}

                tr:hover {{background-color:grey;}}

                .mystyle td {{
                    font-size: 12px;
                    position:relative;
                    padding: 5px;
                    width:10%;
                    color: black;
                border-left: 1px solid #333;
                border-right: 1px solid #333;
                background: rgba(255, 253, 207, 0.58);
                text-align: center;
                }}

                table.mystyle td:first-child {{ text-align: left; }}   

                table.mystyle thead {{
                background: #333333;
                font-size: 14px;
                position:relative;
                border-bottom: 1px solid white;
                border-left: 1px solid white;
                border-right: 1px solid white;
                border-top: 1px solid black;
                }}

                table.mystyle thead th {{
                line-height: 200%;
                font-size: 14px;
                font-weight: normal;
                color: #fffffa;
                text-align: center;
                transition:transform 0.25s ease;
                }}

                table.mystyle thead th:hover {{
                    -webkit-transform:scale(1.01);
                    transform:scale(1.01);
                }}

                table.mystyle thead th:first-child {{
                border-left: none;
                }}

                .topnav {{
                overflow: hidden;
                background-color: black;
                opacity: 0.9;
                }}

                .topnav a {{
                float: right;
                display: block;
                color: #333333;
                text-align: center;
                padding: 12px 15px;
                text-decoration: none;
                font-size: 12px;
                position: relative;
                border: 1px solid #6c3;
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                }}

                #summary{{
                box-shadow: 0 0 80px rgba(200, 112, 1120, 0.4);
                position: relative;
                width:50%;
                cursor: pointer;
                padding: .1%;
                border-style: outset;
                border-radius: 1px;
                border-width: 1px;
                }}
                
                #logo{{
                box-shadow: 0 0 80px rgba(200, 112, 1120, 0.4);
                position: relative;
                cursor: pointer;
                border-style: outset;
                border-radius: 1px;
                border-width: 1px;
                }}

                .topnav a.active {{
                background-color: #333333;
                color: white;
                font-weight: lighter;
                }}

                .topnav .icon {{
                display: none;
                }}

                @media screen and (max-width: 600px) {{
                .topnav a:not(:first-child) {{display: none;}}
                .topnav a.icon {{
                    color: #DBDB40;
                    float: right;
                    display: block;
                }}
                }}

                @media screen and (max-width: 600px) {{
                .topnav.responsive {{position: relative;}}
                .topnav.responsive .icon {{
                    position: absolute;
                    right: 0;
                    top: 0;
                }}
                .topnav.responsive a {{
                    float: none;
                    display: block;
                    text-align: left;
                }}
                }}

                * {{
                box-sizing: border-box;
                }}

                img {{
                vertical-align: middle;
                }}

                .containers {{
                position: relative;
                }}

                .mySlides {{
                display:none;
                width:90%;
                }}

                #slideshow {{
                cursor: pointer;
                margin:.01% auto;
                position: relative;
                width: 70%;
                height: 55%;
                }}

                #ps{{
                height: 10%;
                margin-top: 0%;
                margin-bottom: 90%;
                background-position: center;
                background-repeat: no-repeat;
                background-blend-mode: saturation;
                }}

                #slideshow > div {{
                position: relative;
                width: 90%;
                }}

                #download {{
                background-color: #333333;
                border: none;
                color: white;
                font-size: 12px;
                padding: 13px 20px 15px 20px;
                cursor: pointer;
                }}

                #download:hover {{
                background-color: RoyalBlue;
                }}
                .glow {{
                    font-size: 15px;
                    color: seashell;
                    text-align: center;
                }}
                .reportDiv {{
                    overflow-x: auto;
                    text-align: center;
                }}
                .predictionDiv {{
                    overflow-x: auto;
                    text-align: center;
                    margin-left:4%;
                }}
              
                #report{{
                    box-shadow: 0 0 80px rgba(87, 237, 183, 0.4);
                    overflow-x: auto;
                    min-width:70%;
                }}

                #nestle-section{{
                    float:left;
                    width:100%;
                    position:relative;
                }}

                #nestle-section label{{
                    float:left;
                    width:100%;
                    background:#333;
                    color:#fff;
                    padding:1px 0;
                    text-align:center;
                    cursor:pointer;
                    border:1px solid #818357;
                }}

                #nestle-section label:hover {{background-color:grey;}}

                #nestle-section .tab-content1{{
                    padding:0 10px;
                    height:0;
                    -moz-transition: height 1s ease;
                    -webkit-transition: height 1s ease;
                    -o-transition: height 1s ease;
                    transition: height 1s ease;
                    overflow:hidden;
                }}

                #nestle-section  input:checked + label + .tab-content1{{
                    padding: 10px;
                    height: auto;
                    -moz-transition: height 1s ease;
                    -webkit-transition: height 1s ease;
                    -o-transition: height 1s ease;
                    transition: height 1s ease;
                    overflow: scroll;
                    display: block;
                }}

                #nestle-section input:checked + label{{
                    background:#d5e69a;
                    color:#333;

                }}#nestle-section input{{
                    display:none;
                }}
            </style>
    <body bgcolor="white">
        <div class="reportDiv">""" + "".join(graphs) + """</div>
        <div id="nestle-section">
        <input type="radio" id="tab1" name="tabs1" checked=""/><label for="tab1">Summary Details</label><div class="tab-content1">
        <div class="reportDiv"> """ + execution_summary  + """ alt='execution summary' id='reportDiv' onClick='zoom(this)'></img></br></div></div>
        <input type="radio" id="tab2" name="tabs" checked=""/><label for="tab2">OS Summary</label><div class="tab-content1">
          <div class="reportDiv">""" + monthlyStats + \
          """ </div></div><input type="radio" id="tab3" name="tabs" checked=""/><label for="tab3">Issues</label><div class="tab-content1">
          <div class="reportDiv">""" +issues + \
          """ </div></div><input type="radio" id="tab4" name="tabs" checked=""/><label for="tab4">Custom Failure Reasons</label><div class="tab-content1">
          <div class="reportDiv">""" + failurereasons + \
          """ </div></div><input type="radio" id="tab5" name="tabs" checked=""/><label for="tab5">Top Failed Tests</label><div class="tab-content1">
          <div class="reportDiv">""" +topfailedtable + \
          """ </div>
          </div><input type="radio" id="tab6" name="tabs" checked=""/><label for="tab6">Top Recommendations</label><div class="tab-content1">
          <div class="reportDiv">""" + recommendations + """ </div></div>
          <input type="radio" id="tab7" name="tabs" checked=""/><label for="tab7">Summary</label><div class="tab-content1">
             <div><div class="reportDiv">""" + execution_status + """</div></div></div></body>""")

if __name__ == "__main__":
    start = datetime.now().replace(microsecond=0)
    live_report_filename = "live.html"
    email_report_filename = "email.html"
    try:
        CQL_NAME = str(sys.argv[1])
        OFFLINE_TOKEN = str(sys.argv[2])
    except Exception:
        raise Exception(
            "Pass the mandatory parameters like cloud name and offline token"
        )
    orchestrationIssues = ["already in use"]
    labIssues = ["HANDSET_ERROR"]
    REPORTING_SERVER_URL = "https://" + CQL_NAME + ".reporting.perfectomobile.com"
    api_url = REPORTING_SERVER_URL + "/export/api/v1/test-executions"
    resources = []
    topTCFailureDict = {}
    topDeviceFailureDict = {}

    # report = "report|jobName=test|jobNumber=1|startDate=123|endDate=1223|consolidate=/Users/temp|xlformat=csv|PORT=8888"
    report = sys.argv[3]
    try:
        criteria = ""
        jobName = ""
        jobNumber = ""
        startDate = ""
        endDate = ""
        consolidate = ""
        xlformat = "csv"
        port = ""
        temp = ""
        report_array = report.split("|")
        for item in report_array:
            if "jobName" in item: jobName, criteria =  get_report_details(item, temp, "jobName", criteria)
            if "jobNumber" in item: jobNumber, criteria =  get_report_details(item, temp, "jobNumber", criteria)
            if "startDate" in item: startDate, criteria =  get_report_details(item, temp, "startDate", criteria)
            if "endDate" in item: endDate, criteria =  get_report_details(item, temp, "endDate", criteria)
            if "consolidate" in item: consolidate, criteria =  get_report_details(item, temp, "consolidate", criteria)
            if "xlformat" in item: xlformat, criteria =  get_report_details(item, temp, "xlformat", criteria)
            if "port" in item: port, criteria =  get_report_details(item, temp, "port", criteria)
    except Exception as e:
        raise Exception( "Verify parameters of report, split them by | seperator" + str(e) )
        sys.exit(-1)
    filelist = glob.glob(os.path.join("*." + xlformat))
    for f in filelist:
        os.remove(f)
    filelist = glob.glob(os.path.join("*_failures.txt" ))
    for f in filelist:
        os.remove(f)
    filelist = glob.glob(os.path.join("*.html" ))
    for f in filelist:
        os.remove(f)

    graphs, df = prepareReport(jobName, jobNumber)
    if not jobName:
        criteria = "start: "  + startDate + "; end: " + endDate
    if consolidate != "":
        criteria = "startTime: "  + str(min(df['startTime'])) + "; endTime: " + str(max(df['startTime'])) + "; consolidate: " + consolidate
        if jobName !="":
            criteria += "; jobName:" + jobName
        if jobNumber !="":
            criteria += "; jobNumber:" + jobNumber
    execution_summary = create_summary(df, CQL_NAME.upper() + " Summary Report for " + criteria, "status", "device_summary")
    failed = df[(df['status'] == "FAILED")]
    passed = df[(df['status'] == "PASSED")]
    blocked = df[(df['status'] == "BLOCKED")]
    failed_blocked = df[(df['status'] == "FAILED") | (df['status'] == "BLOCKED")]
    totalUnknownCount = df[(df['status'] == "UNKNOWN")].shape[0]
    totalTCCount = df.shape[0]
    #monthly stats
    df['platforms/0/deviceType'] = df['platforms/0/deviceType'].fillna('Others')
    df['platforms/0/os'] = df['platforms/0/os'].fillna('Others')
    df = df.rename(columns={'platforms/0/deviceType': 'Platform', 'platforms/0/os' : 'OS', 'status' : 'Test Status', 'failureReasonName' : 'Custom Failure Reason'})
    monthlyStats = df.pivot_table(index = ["month",  "week", "Platform", "OS"], 
                columns = "Test Status" , 
                values = "name", 
                aggfunc = "count", margins=True, fill_value=0)\
            .fillna('')
    for column in monthlyStats.columns:
        monthlyStats[column] = monthlyStats[column].astype(str).replace('\.0', '', regex=True)
    monthlyStats = monthlyStats.to_html( classes="mystyle", table_id="report", index=True, render_links=True, escape=False ).replace('<tr>', '<tr align="center">')
    failurereasons = pandas.crosstab(df['Custom Failure Reason'],df['Test Status'])
    # print (failurereasons)
    failurereasons = failurereasons.to_html( classes="mystyle", table_id="report", index=True, render_links=True, escape=False )
    #top failed TCs
    topfailedTCNames = failed.groupby(['name']).size().reset_index(name='#Failed').sort_values('#Failed', ascending=False).head(5)
    reportURLs = []
    for ind in topfailedTCNames.index:
        reportURLs.append(failed.loc[failed['name'] == topfailedTCNames['name'][ind], 'reportURL'].iloc[0])
    topfailedTCNames['Result'] = reportURLs
    topfailedTCNames['Result'] = topfailedTCNames['Result'].apply(lambda x: '{0}'.format(x))
    for ind in topfailedTCNames.index:
        topfailedTCNames.loc[topfailedTCNames['name'].index == ind, 'name']  = '<a target="_blank" href="' + topfailedTCNames['Result'][ind] + '">' + topfailedTCNames['name'][ind] + '</a>'
    topfailedTCNames = topfailedTCNames.drop('Result', 1)
    topfailedTCNames.columns = ['Top 5 Failed Tests', '#Failed']
    # print(str(topfailedTCNames))
    topfailedtable = topfailedTCNames.to_html( classes="mystyle", table_id="report", index=False, render_links=True, escape=False )

    #recommendations
    orchestrationIssues = ["already in use"]
    labIssues = ["HANDSET_ERROR", "ERROR: No device was found"]
    regEx_Filter = "Build info:|For documentation on this error|at org.xframium.page|Scenario Steps:| at WebDriverError|\(Session info:|XCTestOutputBarrier\d+|\s\tat [A-Za-z]+.[A-Za-z]+.|View Hierarchy:|Got: |Stack Trace:|Report Link|at dalvik.system|Output:\nUsage|t.*Requesting snapshot of accessibility"
    labIssuesCount = 0
    scriptingIssuesCount = 0
    orchestrationIssuesCount = 0
    cleanedFailureList = {}
    suggesstionsDict = {}
    totalFailCount = failed.shape[0]
    totalPassCount = passed.shape[0]
    blockedCount = blocked.shape[0]
    # failures count
    failuresmessage = failed_blocked.groupby(['message']).size().reset_index(name='#Failed').sort_values('#Failed', ascending=False)

    for commonError, commonErrorCount in failuresmessage.itertuples(index=False):
        for labIssue in labIssues:
            if re.search(labIssue, commonError):
                labIssuesCount += commonErrorCount
                break
        for orchestrationIssue in orchestrationIssues:
            if re.search(orchestrationIssue, commonError):
                orchestrationIssuesCount += commonErrorCount
                break
        error = commonError
        regEx_Filter = "Build info:|For documentation on this error|at org.xframium.page|Scenario Steps:| at WebDriverError|\(Session info:|XCTestOutputBarrier\d+|\s\tat [A-Za-z]+.[A-Za-z]+.|View Hierarchy:|Got: |Stack Trace:|Report Link|at dalvik.system|Output:\nUsage|t.*Requesting snapshot of accessibility"
        if re.search(regEx_Filter, error):
            error = str(re.compile(regEx_Filter).split(error)[0])
            if "An error occurred. Stack Trace:" in error:
                error = error.split("An error occurred. Stack Trace:")[1]
        if re.search("error: \-\[|Fatal error:", error):
            error = str(re.compile("error: \-\[|Fatal error:").split(error)[1])
        if error.strip() in cleanedFailureList:
            cleanedFailureList[error.strip()] += 1
        else:
            cleanedFailureList[error.strip()] = commonErrorCount
        scriptingIssuesCount = (totalFailCount + blockedCount) - (orchestrationIssuesCount + labIssuesCount)

    # Top 5 failure reasons
    topFailureDict = {}

    failureDict = Counter(cleanedFailureList)
    for commonError, commonErrorCount in failureDict.most_common(5):
        topFailureDict[commonError] = int(commonErrorCount)

    # reach top errors and clean them
    i = 0
    for commonError, commonErrorCount in topFailureDict.items():
        if "ERROR: No device was found" in commonError:
            error = (
                "Raise a support case as the error: *|*"
                + commonError.strip()
                + "*|* occurs in *|*"
                + str(commonErrorCount)
                + "*|* occurrences"
            )
        elif "Cannot open device" in commonError:
            error = (
                "Reserve the device/ use perfecto lab auto selection feature to avoid the error:  *|*"
                + commonError.strip()
                + "*|* occurs in *|*"
                + str(commonErrorCount)
                + "*|* occurrences"
            )
        elif '(UnknownError) Failed to execute command button-text click: Needle not found for expected value: "Allow" (java.lang.RuntimeException)' in commonError:
            error = (
            "Allow text/popup was not displayed as expected. It could be an environment issue as the error: *|*"
            + commonError.strip()
            + "*|* occurs in *|*"
            + str(commonErrorCount)
            + "*|* occurrences"
        )
        else:
            error = (
                "Fix the error: *|*"
                + commonError.strip()
                + "*|* as it occurs in *|*"
                + str(commonErrorCount)
                + "*|* occurrences"
            )
        suggesstionsDict[error] = commonErrorCount
    eDict = edict(
            {
                "status": [
                    {
                    "#Total": "Count ->",
                    "#Executions": totalTCCount,
                    "#Pass" : totalPassCount,
                    "#Failed" : totalFailCount,
                    "#Blocked" : blockedCount,
                    "#Unknowns": totalUnknownCount,
                    "Overall Pass %": str(int(percentageCalculator(totalPassCount, totalTCCount))) + "%",
                    },
                ],
                "issues": [
                {
                    "#Issues": "Count ->",
                    "#Scripting": scriptingIssuesCount,
                    "#Lab": labIssuesCount,
                    "#Orchestration": orchestrationIssuesCount,
                    },
                ],
                "recommendation": [
                    {
                    "Recommendations": "-",
                    "Rank": 1,
                        "impact": "0",
                    },
                    {
                    "Recommendations": "-",
                        "Rank": 2,
                        "impact": "0",
                    },
                    {
                        "Recommendations": "-",
                        "Rank": 3,
                        "impact": "0",
                    },
                    {
                        "Recommendations": "-",
                        "Rank": 4,
                        "impact": "0",
                    },
                    {
                        "Recommendations": "-",
                        "Rank": 5,
                        "impact": "0",
                    },
                ],
            }
        )
    jsonObj = edict(eDict)
    if float(percentageCalculator(totalUnknownCount, totalTCCount)) >= 30:
        suggesstionsDict[
            "# Fix the unknowns. The unknown script ratio is too high (%) : "
            + str(percentageCalculator(totalUnknownCount, totalTCCount))
            + "%"
        ] = percentageCalculator(
            totalPassCount + totalUnknownCount, totalTCCount
        ) - percentageCalculator(
            totalPassCount, totalTCCount
        )
    if len(suggesstionsDict) < 5:
        if (topfailedTCNames.shape[0]) > 1:
            for tcName, status in topfailedTCNames.itertuples(index=False):
                suggesstionsDict[
                    "# Fix the top failing test: "
                    + tcName
                    + " as the failures count is: "
                    + str(int((str(status).split(",")[0]).replace("[", "").strip()))
                ] = 1
                break

    if len(suggesstionsDict) < 5:
        if int(percentageCalculator(totalFailCount, totalTCCount)) > 15:
            if totalTCCount > 0:
                suggesstionsDict[
                    "# Fix the failures. The total failures % is too high (%) : "
                    + str(percentageCalculator(totalFailCount, totalTCCount))
                    + "%"
                ] = totalFailCount
    if len(suggesstionsDict) < 5:
        if float(percentageCalculator(totalPassCount, totalTCCount)) < 80 and (
            totalTCCount > 0
        ):
            suggesstionsDict[
                "# Fix the failures. The total pass %  is too less (%) : "
                + str(int(percentageCalculator(totalPassCount, totalTCCount)))
                + "%"
            ] = (
                100
                - (
                    percentageCalculator(
                        totalPassCount + totalUnknownCount, totalTCCount
                    )
                    - percentageCalculator(totalPassCount, totalTCCount)
                )
            ) - int(
                percentageCalculator(totalPassCount, totalTCCount)
            )
    if len(suggesstionsDict) < 5:
        if totalTCCount == 0:
            suggesstionsDict[
                "# There are no executions for today. Try Continuous Integration with any tools like Jenkins and schedule your jobs today. Please reach out to Professional Services team of Perfecto for any assistance :) !"
            ] = 100
        elif int(percentageCalculator(totalPassCount, totalTCCount)) > 80:
            print(str(int(percentageCalculator(totalPassCount, totalTCCount))))
            suggesstionsDict["# Great automation progress. Keep it up!"] = 0

        int(percentageCalculator(totalFailCount, totalTCCount)) > 15
    topSuggesstionsDict = Counter(suggesstionsDict)
    counter = 0
    totalImpact = 0
    for sugg, commonErrorCount in topSuggesstionsDict.most_common(5):
        impact = 1
        if sugg.startswith("# "):
            sugg = sugg.replace("# ", "")
            impact = commonErrorCount
        else:
            impact = percentageCalculator(
                totalPassCount + commonErrorCount, totalTCCount
            ) - percentageCalculator(totalPassCount, totalTCCount)
        jsonObj.recommendation[counter].impact = str(("%.2f" % round(impact, 2))) + "%"
        jsonObj.recommendation[counter].Recommendations = (
            html.escape(sugg.replace("*|*", "'").replace("{","{{").replace("}","}}").strip())
        )
        totalImpact += round(impact, 2)
        counter += 1
    execution_status = pandas.DataFrame.from_dict(jsonObj.status)
    execution_status = execution_status.to_html( classes="mystyle", table_id="report", index=False, render_links=True, escape=False )
    issues = pandas.DataFrame.from_dict(jsonObj.issues)
    issues = issues.to_html( classes="mystyle", table_id="report", index=False, render_links=True, escape=False )
    recommendations = pandas.DataFrame.from_dict(jsonObj.recommendation)
    if totalImpact > 100:
        recommendations.columns = ['Recommendations', 'Rank', 'Impact to Pass %']
    else:
        recommendations.columns = ['Recommendations', 'Rank', 'Impact to Pass % [Total - ' + str(round(totalImpact,2)) + '%]']
    recommendations = recommendations[recommendations.Recommendations.astype(str) != "-"]
    recommendations = recommendations.to_html( classes="mystyle", table_id="report", index=False, render_links=True, escape=False )
    
    with open(email_report_filename, "a") as f:
        f.write(get_html_string(graphs).format(table=df.to_html( classes="mystyle", table_id="report", index=False , render_links=True, escape=False)))

    graphs.clear()

    with open(live_report_filename, "a") as f:
        f.write(get_html_string(graphs).format(table=df.to_html( classes="mystyle", table_id="report", index=False , render_links=True, escape=False)))

    end = datetime.now().replace(microsecond=0)
    print("Total Time taken:" + str(end - start))
    import http.server
    import socketserver
    import socket
    import webbrowser
    from psutil import process_iter
    from signal import SIGTERM # or SIGKILL

    if port != "":
        PORT = int(port)
        try:
            for proc in process_iter():
                for conns in proc.connections(kind='inet'):
                    if conns.laddr.port == PORT:
                        proc.send_signal(SIGTERM) # or SIGKILL
        except:
            pass
        Handler = http.server.SimpleHTTPRequestHandler
        url = "http://" + socket.gethostbyname(socket.gethostname()) + ":" + str(PORT) + "/" + live_report_filename
        print("Live dashboard url: " + url)
        with socketserver.TCPServer(("", PORT), Handler) as httpd:
            print("serving at port", PORT)
            httpd.serve_forever()

    