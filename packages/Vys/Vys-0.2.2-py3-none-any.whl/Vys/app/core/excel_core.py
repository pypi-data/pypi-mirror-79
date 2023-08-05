#! python3
import openpyxl
import os


class ExcelCore:
    def __init__(self):
        self.data = {}

    def key_sheet(self, workbook, read_only=True):  # TODO This doesnt work, please fix
        wb = openpyxl.load_workbook(workbook, read_only=read_only)
        if len(wb.sheetnames) != 1:
            choice = False
            while choice not in wb.sheetnames:
                print('Type the name of the sheet you wish to use:', end='')
                for sheet in wb.sheetnames:
                    print(' %s,' % sheet, end='')
                print('')
                choice = input()
            sheet = wb[choice]
        else:
            sheet = wb[wb.sheetnames[0]]
        header_data = tuple()
        bulk_data = list()
        iteration = 1
        for row in sheet.values:
            if iteration == 1:
                header_data = row
                iteration += 1
            else:
                bulk_data.append(row)

        self._key_data(bulk_data, header_data, sheet=sheet)

    def key_workbook(self, workbook, read_only=True):
        wb = openpyxl.load_workbook(workbook, read_only=read_only)
        for sheet in wb.sheetnames:
            sheet_data = list()
            header_data = tuple()
            ws = wb[sheet]
            iteration = 1
            for row in ws.values:
                if iteration == 1:
                    header_data = row
                    iteration += 1
                else:
                    sheet_data.append(row)
            self._key_data(sheet_data, header_data, sheet=sheet)

    def create_workbook(self, data, headers):
        wb = openpyxl.Workbook()
        for i in data:
            wb.create_sheet(i)
            ws = wb[i]
            col = 1
            for f in headers:
                ws.cell(row=1, column=col, value=f)
                row = 1
                for x in data[i]:
                    row += 1
                    ws.cell(row=row, column=col, value=x[f])
                col += 1
        del wb['Sheet']
        wb.save('output.xlsx')
        os.startfile('output.xlsx')

    def _key_data(self, bulk_data, header_data, sheet='Sheet1'):
        print(sheet)
        keyed_data = []
        for customer_row in bulk_data:
            c = {}
            current_row = 0
            for item in customer_row:
                c[header_data[current_row]] = item
                current_row += 1
            keyed_data.append(c)
        self.data[sheet] = keyed_data
