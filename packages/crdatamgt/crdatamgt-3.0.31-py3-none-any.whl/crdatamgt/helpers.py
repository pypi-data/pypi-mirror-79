import glob
import os
import re

import openpyxl as ox
import pandas as pd
import simplelogging
from xloaderx.loading import loading


def write_yaml():
    # 'PROJECT DIRECTORY': None,
    return {
        "FORMULATION DIRECTORY": None,
        "RESULTS DIRECTORY": None,
    }


def data_start(sheet, header=False):
    y = sheet.values
    _header = []
    for x in y:
        if header:
            _header.append(x)
        if x[0] == "67 6f":
            return _header, y, True
        if x[0] == "6e 74":
            return _header, y, False
    return _header, y, False


def data_extraction(wb, sheet, use_header=False):
    dataframe = pd.DataFrame()
    header = pd.DataFrame()
    if sheet in wb.sheetnames:
        data_wb = wb[sheet]
        head_data, data, transpose = data_start(data_wb, use_header)
        if transpose:
            data_transposed = [
                transposed for transposed in zip(*[line for line in data if line[0]])
            ]
            dataframe, header = _extraction(
                data_transposed, head_data, header, dataframe, use_header
            )
        else:
            data_ = [line for line in data]
            dataframe, header = _extraction(
                data_, head_data, header, dataframe, use_header
            )
            dataframe = dataframe[dataframe.columns.dropna()]
    return [dataframe, [header, dataframe]][use_header]


def _extraction(data, head_data, header, dataframe, use_header):
    if data:
        dataframe = pd.DataFrame(columns=data[0], data=[line for line in data[1:]])
        dataframe = dataframe.truncate(after=dataframe.last_valid_index())
    if use_header:
        head_data = [line for line in head_data]
        if head_data:
            header = pd.DataFrame(data=[line for line in head_data]).dropna(
                axis=1, how="all"
            )
    return dataframe, header


def workbook_save(name, path, dataframe, sheet, **kwargs):
    data_path = os.path.join(path, name + ".xlsx")
    writer = pd.ExcelWriter(data_path, engine="xlsxwriter")
    dataframe.to_excel(writer, sheet_name=sheet, startrow=1, header=False, index=False)
    if kwargs.get("header"):
        workbook = writer.book
        header_format = workbook.add_format(kwargs.get("header"))
        for _sheet in writer.sheets:
            worksheet = writer.sheets[_sheet]
            for col_num, value in enumerate(dataframe.columns.values):
                worksheet.write(0, col_num, value, header_format)
    writer.save()


def workbook_load_file(path, write=False) -> [ox.workbook]:
    results_wb = []
    excels = results_excels(path)
    loading_indictator = loading(len(excels), additional_text="Reading workbooks")
    for e in excels:
        next(loading_indictator)
        e_path = os.path.join(path, e)
        if write:
            wb = ox.load_workbook(filename=e_path, data_only=True)
        else:
            wb = ox.load_workbook(filename=e_path, data_only=True, read_only=True)
        results_wb.append(wb)
    return results_wb


def workbook_load_path(path, write=False):
    data_file = [x for x in os.listdir(path) if ".xlsx" in os.path.splitext(x)]
    if len(data_file) > 1:
        # TODO: Logging
        pass
    else:
        pass
    data_file = data_file[0]
    data_path = os.path.join(path, data_file)
    if write:
        wb = ox.load_workbook(filename=data_path, data_only=True)
    else:
        wb = ox.load_workbook(filename=data_path, data_only=True, read_only=True)
    return wb


def rename_dictionary():
    return {
        "pg": "propylene glycol",
        "propylen glycol": 'propylene glycol"',
        "glycerine": "glycerin",
    }


def topic_directories(path) -> (list, list):
    topics = [topic for topic in os.listdir(path) if re.search(r"^Topic", topic)]

    directories = [
        os.path.join(path, topic)
        for topic in topics
        if os.path.isdir(os.path.join(path, topic))
    ]

    return topics, directories


def is_dir(path):
    path_list = [
        globed
        for globed in glob.glob(r"{}".format(path + "/*"))
        if os.path.isdir(globed)
    ]
    return path_list


def results_excels(path) -> list:
    # Re to filter out temporary hidden files and junk files
    results = [
        excel
        for excel in os.listdir(path)
        if os.path.splitext(excel)[-1].lower() == ".xlsx"
        and not re.search(r"^(~\$)", os.path.splitext(excel)[0])
        and not os.stat(os.path.join(path, excel)).st_size <= 1
    ]
    return results


def log_format(formatting=None):
    if not formatting:
        formatting = "simple"
    if formatting == "simple":
        return "%(asctime)20s [%(levelname)-1s]" "%(message)s"


def logger(name=None):
    return simplelogging.get_logger(
        name=name,
        file_name="Error Logging.log",
        console_format=log_format(),
        console_level=simplelogging.WARNING,
        file_format=log_format(),
        file_level=simplelogging.DEBUG,
    )

def exception():
    pass