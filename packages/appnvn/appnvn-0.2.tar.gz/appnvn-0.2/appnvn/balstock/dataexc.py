import pandas as pd
from shutil import copyfile
import openpyxl
import openpyxl.styles
from openpyxl.styles import PatternFill
class comparetwofile():
    def __init__(self,path_OLD = None, 
                path_NEW = None, 
                index_col = None,
                sheetname = 0,
                usernamein = None,
                dt = None,
                pathtcsvtosavedata = None,
                pathtorgindiff = None,
                difpathtobk = None
                ) :
        self.path_OLD = path_OLD
        self.path_NEW = path_NEW
        self.index_col = index_col
        self.sheetname = sheetname
        self.usernamein = usernamein
        self.difpathtobk = difpathtobk
        self.dt = dt
        self.pathtcsvtosavedata = pathtcsvtosavedata
        self.pathtorgindiff = pathtorgindiff
    def excel_diff(self):
        # get sheet name file 
        xl = pd.ExcelFile(self.path_OLD)

        shname =  xl.sheet_names[0]  # see all sheet names

        df_OLD = pd.read_excel(self.path_OLD,sheet_name=self.sheetname,
                                index_col=self.index_col).fillna("")
        df_NEW = pd.read_excel(self.path_NEW,sheet_name=self.sheetname, 
                                index_col=self.index_col).fillna("")
        # Perform Diff
        dfDiff = df_NEW.copy()
        droppedRows = []
        newRows = []
        diffRows = []
        rowandcolumn = []
        indexoldnew = []
        cols_OLD = df_OLD.columns
        cols_NEW = df_NEW.columns
        sharedCols = list(set(cols_OLD).intersection(cols_NEW))
        for row in dfDiff.index:
            if (row in df_OLD.index) and (row in df_NEW.index):
                for col in sharedCols:
                    value_OLD = df_OLD.loc[row,col]
                    value_NEW = df_NEW.loc[row,col]
                    if value_OLD==value_NEW:
                        dfDiff.loc[row,col] = df_NEW.loc[row,col]
                    else:
                        dfDiff.loc[row,col] = ('{}→{}').format(value_OLD,
                                                                value_NEW)

                        dfDiff.loc[row,100] = ('{} changed data to {} at {}').format(self.usernamein,
                                                                                value_NEW,self.dt)
                        # get index column
                        col_index = dfDiff.columns.get_loc(col)

                        rowandcolumn.append([row + 1,
                                            col_index])

                        indexoldnew.append([row + 1,col_index,value_OLD,value_NEW])

                        diffRows.append(row)
            else:
                newRows.append(row)
        # save as to csv file 
        datachange = [[self.usernamein,
                    self.dt,indexoldnew]]
        df = pd.DataFrame(data=datachange)
        df.to_csv (self.pathtcsvtosavedata,
                    index = None,
                    header=False, 
                    mode = "a")    
        #wcsv1.writefilecsvFromRowArr()        
        for row in df_OLD.index:
            if row not in df_NEW.index:
                droppedRows.append(row)
                dfDiff = dfDiff.append(df_OLD.loc[row,:])

        dfDiff = dfDiff.sort_index().fillna('')
        # Save output and format
        #fname = "Test12.xlsx"
        copyfile(self.path_NEW,self.path_OLD)
        copyfile(self.path_NEW,self.difpathtobk)
        #Using openpyxl to fill collor 

        wb = openpyxl.load_workbook(self.difpathtobk)
        sheets = wb.sheetnames


        ws = wb[sheets[0]]
        for cell in rowandcolumn: 
            ws.cell(row=int(cell[0]) + 1, column=int(cell[1]) + 1).fill = PatternFill('solid',openpyxl.styles.colors.GREEN)
        wb.save(self.difpathtobk)

        writer = pd.ExcelWriter(self.pathtorgindiff, 
                                engine='xlsxwriter')


        #compare 2 file 
        dfDiff.to_excel(writer, sheet_name='DIFF', 
                                index=False)

        diffRows = list(set(diffRows+newRows+droppedRows))

        # get xlsxwriter objects
        workbook  = writer.book
        worksheet = writer.sheets['DIFF']
        #worksheet_org = writerbk.sheets["TON15122019"]
        worksheet.hide_gridlines(2)
        worksheet.set_default_row(15)

        # define formats
        date_fmt = workbook.add_format({'align': 'center', 'num_format': 'yyyy-mm-dd'})
        center_fmt = workbook.add_format({'align': 'center'})
        number_fmt = workbook.add_format({'align': 'center', 'num_format': '#,##0.00'})
        cur_fmt = workbook.add_format({'align': 'center', 'num_format': '$#,##0.00'})
        perc_fmt = workbook.add_format({'align': 'center', 'num_format': '0%'})
        grey_fmt = workbook.add_format({'font_color': '#E0E0E0'})
        highlight_fmt = workbook.add_format({'font_color': '#FF0000', 'bg_color':'#B1B3B3'})
        new_fmt = workbook.add_format({'bg_color': '#C6EFCE','font_color': '#32CD32','bold':True})

        # set format over range
        ## highlight changed cells
        worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
                                                'criteria': 'containing',
                                                'value':'→',
                                                'format': highlight_fmt})
                                    
        # highlight new/changed rows
        for row in range(dfDiff.shape[0]):
            if row+1 in newRows:
                worksheet.set_row(row+1, 15, new_fmt)
            if row+1 in droppedRows:
                worksheet.set_row(row+1, 15, grey_fmt)
        writer.save()