import pandas as pd
import os
import sys
from openpyxl import load_workbook
from pynvn.path import (IsRunningInPyinstallerBundle,
                        resource_path_is_from_pyinstall_and_dev,
                        ExtractFileNameFromPath,
                        PathFromFileNameAndDirpath,
                        )
                    
from pynvn.dict import credict
from pynvn.excel import toexcel
from pynvn.data import grouper,duprowdata
from pynvn.dict import updictjoint
def celllist (  sheet,
                minrow,
                mincol, 
                maxcol, 
                maxrow):
    ArrRafterTT = []
    for row in sheet.iter_cols (min_row = minrow,
                                min_col = mincol,
                                max_col = maxcol, 
                                max_row = maxrow):
        ArrRafter = [cell.value for cell in row]
        ArrRafterTT.append(ArrRafter)
    return ArrRafterTT
def ReturnCountRafter (kk,sheet):
        last_row = 0
        while True:
            cell_value = sheet.cell(row = kk, 
                                    column = 1).value
            if cell_value is not None:              
                last_row += 1
                kk = kk + 1
            else:
                break
        return last_row
# check running in Pyintaller or not ?
if IsRunningInPyinstallerBundle():
    DataExcel = resource_path_is_from_pyinstall_and_dev(FileName = 'DataALL - Template.xlsx',
                                                         Subfolder="Data",
                                                         Is_Directory_Path_To_SubFolder= True,
                                                         dir_path=sys._MEIPASS)
    
def CreateFileCSV(pathin,pathout):
    filename = "Config_Setting.csv"
    fullpath = os.path.join(pathin,filename)
    pathexc = os.path.join(pathout,"new_big_file.xlsx") 

    # get path store data to handling
    Right_Genneral_All_path = PathFromFileNameAndDirpath(dir_path =pathin,
                                                         filename ="Right_Genneral_All.csv"
                                                         )
                                    
    Left_Genneral_All_path = PathFromFileNameAndDirpath(dir_path =pathin,
                                                         filename ="Left_Genneral_All.csv"
                                                         )
                                
    # key value 
    keyvalue = ["ValueGeneral",
                "Columnmove",
                "LocationCellForMoveColumn",
                "GenneralColumnNotChange",
                "GeneralConcernRaffter",
                "Genneral_Select",
                "LocationOfRowLeft",
                "LocationOfRowRight",
                "ExcelCellForMoveColumnRight",
                "LocationOfPurlin",
                "startrow",
                "Index_Path_From_CSV",
                "TitleLocationMoveColumn",
                ]
    
    #location in conf
    locvalue = [12,16,17,19,20,21,23,24,25,26,30,28,27]
    #create dict
    credict_c = credict(KeyValues = keyvalue,
                        LocConf = locvalue,
                        Config_Setting_Path = fullpath)

    credict_list = credict_c.Dictfromkeyandvalueconf()

    # create arr from keyvalue  
    valgen = credict_list.get("ValueGeneral", "")
    colmv = credict_list.get("Columnmove", "")
    locmvcol = tuple(credict_list.get("LocationCellForMoveColumn", ""))
    gencolnotchg = credict_list.get("GenneralColumnNotChange", "")
    genconraf = credict_list.get("GeneralConcernRaffter", "")
    gensel = credict_list.get("Genneral_Select", "")
    locrowleft = credict_list.get("LocationOfRowLeft", "")
    locrowright = credict_list.get("LocationOfRowRight", "")
    excemvcolright = credict_list.get("ExcelCellForMoveColumnRight", "")
    locpur = credict_list.get("LocationOfPurlin", "")
    strow = credict_list.get("startrow", "")
    pathindexcsv = credict_list.get("Index_Path_From_CSV", "")
    titlelocmovecol = credict_list.get("TitleLocationMoveColumn", "")
    book = load_workbook(pathexc)
    #write csv 
    sheet = book.get_sheet_by_name('General Member')

    #Create dict Genneral value 
    cell_obj_Arr = celllist(sheet,strow[0] \
             + 1,1,len(valgen),strow[0] + 2
                            )

    dictvalgen = dict(zip(valgen,
                    cell_obj_Arr))
    
    pathin = [Left_Genneral_All_path,
            Right_Genneral_All_path]

    for path in pathin:
        # create frame to excel 
        excellframe = toexcel(worksheet = sheet,
                                        path = path,
                                        path_conf =fullpath,
                                        lpath=Left_Genneral_All_path,
                                        rpath = Right_Genneral_All_path)

        # check path 
        if path == Left_Genneral_All_path:
            LocationOfRow = locrowleft
            Pathdt = 'Left.csv'
            loccell = excellframe.reloccol()
            loccellre = loccell
        else:
            LocationOfRow = locrowright
            Pathdt = 'Right.csv'
            loccell = excellframe.reloccol()
            loccellre = loccell

        #create dict for Genneral select:
        cell_obj_Arr = celllist(sheet,int(LocationOfRow[0]) + 1,
                            1,len(gensel),int(LocationOfRow[0]) + 2
                                )
        dictvalgensel = dict(zip(gensel,cell_obj_Arr))

        #create dict for column Selected:
        cell_obj_Arr = celllist(sheet,int(LocationOfRow[1]) + 1,
                    1,len(gencolnotchg),int(LocationOfRow[1]) + 2)
        Dict_Column_Select = dict(zip(gencolnotchg,cell_obj_Arr))

        #create dict for Rafter Selected:
        count = ReturnCountRafter(int(LocationOfRow[2]) + 1,sheet)

        ArrRafterTT = celllist(sheet,int(LocationOfRow[2]) + 1,
                                        1,6,int(LocationOfRow[2]) + count
                                         )

        Dict_Rafter_Select = dict(zip(genconraf,
                                    ArrRafterTT)
                                    )
        #create purlin list

        Arr_Purlin = celllist(sheet,int(LocationOfRow[3]) + 1,1,len(locpur),
                                                int(LocationOfRow[3]) + 2)
        Dict_Purlin = dict(zip(locpur,Arr_Purlin))

        #write path to csv from excel 
        Path_Value = sheet.cell(row = loccellre[0],
                                column = int(loccellre[1]) + 1).value

        Path_Title = sheet.cell(row = loccellre[0],
                             column = loccellre[1]).value
                                 
        Path_Arr = [Path_Title,
                    Path_Value]

        distlistpath = {pathindexcsv[0]:Path_Arr}
        #create group for arr
        locmvcol1 =  list(grouper(2,locmvcol))
        #create dict for move column         
        cell_obj_Arr = [sheet.cell(row =  cell[0],
                        column = cell[1]).value for cell in locmvcol1]

        Title_Arr = titlelocmovecol
        cell_obj_Arr = list(zip (Title_Arr,
                            cell_obj_Arr))
        Dict_Column_Move = dict(zip(colmv,
                            cell_obj_Arr))
        #update Dicts to a gennenral Dict 
        Genenral_Dict = {}

        ListDist = [Dict_Column_Move,
                    Dict_Rafter_Select,\
                        Dict_Column_Select,
                        dictvalgen,
                        dictvalgensel,
                        Dict_Purlin,
                        distlistpath
                    ]

        Genenral_Dict_Sorted = updictjoint(*ListDist)

        df = pd.DataFrame(dict([ (k,pd.Series(v)) for
                                 k,v in Genenral_Dict_Sorted.items()]))

        df_ed = duprowdata(df)
        path = os.path.join(pathout,Pathdt) 
        df_ed.to_csv(path,
                    index=False,
                    header= False)
