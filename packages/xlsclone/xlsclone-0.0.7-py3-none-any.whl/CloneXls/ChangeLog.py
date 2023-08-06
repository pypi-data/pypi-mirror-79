from openpyxl import load_workbook
from datetime import datetime
from copy import copy
import logging

logger = logging.getLogger("pyC")


class ChangeLog:
    def __init__(
            self, source_file, clone_file, log_sheet_name="Change Logs", ignore_headers=[], ignore_sheets=[],
            case_sensitive_ignore=False
    ):
        """

        :param source_file: Path of excel file.
        :param clone_file: Path of clone version of source file.
        :param log_sheet_name: (optional) (default="Change Logs") Name of sheet containing change logs
        :param ignore_headers: (optional) list of headers to be ignored for change log
        :param ignore_sheets: (optional) list of sheet names to be ignored for change log
        :param case_sensitive_ignore: (default=False) If True, ignore sheets & headers list's data will be matched with same case.
        """
        self.source_file = source_file
        self.clone_file = clone_file
        self.log_sheet_name = log_sheet_name
        self.ignore_headers = ignore_headers
        self.ignore_sheets = ignore_sheets
        self.ignore_sheets.append(self.log_sheet_name)
        self.case_senstive_ignore = case_sensitive_ignore
        self.update_ignore_list()
        self.headers_column = {}

    def update_ignore_list(self):
        # Making data inside igonre lists in lower only if case_sensitive_ignore is false, so that it will be later
        # used at the time of comparision
        logger.debug("Updating ignore lists")
        if self.case_senstive_ignore:
            return None
        if self.ignore_headers:
            self.ignore_headers = [header.lower() for header in self.ignore_headers]
            logger.debug("Headers ignore list updated.")
        if self.ignore_sheets:
            self.ignore_sheets = [sheet.lower() for sheet in self.ignore_sheets]
            logger.debug("Sheets ignore list updated.")

    def xlsxChangeLog(self):
        """
        Method to trigger all functionalities including checking changes and updating clone file
        :return:
        """
        # Main method which will do all things including check and updating logs
        source = load_workbook(self.source_file)
        clone = load_workbook(self.clone_file)
        source_sheets = source.get_sheet_names()
        clone_sheets = clone.get_sheet_names()
        self.check_log_sheet(clone_sheets, clone)
        sheets = self.get_sheets(source_sheets, clone_sheets)
        logs = self.get_change_logs(sheets, source, clone)
        if logs:  # if their is no update in the non ignored data than file won't be updated
            self.update_change_logs(logs, clone)
            self.update_clone_file(source, clone, source_sheets)
            clone.save(self.clone_file)
        clone.close()
        source.close()

    def check_log_sheet(self, clone_sheets, clone):
        # Checks if log worksheet is inside clone file or not, if not then it will create one
        if self.log_sheet_name not in clone_sheets:
            sht = clone.create_sheet(self.log_sheet_name)
            headers = ["Date & Time", "Sheet Name", "Header Name", "Row Number", "Old Value", "New Value"]
            for i in range(1, len(headers)+1):  # looping through all headers and inserting them on correct position
                sht.cell(1, i).value = headers[i-1]

    def get_sheets(self, source_sheets, clone_sheets):
        # method is used to get sheets which are not inside ignore list, also deals with case_sensitive functionality
        sheets = []
        for sheet in source_sheets:  # looping through sheets in source file
            if sheet in clone_sheets:  # if sheet is not already present in clone file than change log is skipped
                if self.ignore_sheets:  # if ignore_sheets doesn't have any values than sheet name is directly added in list
                    if self.case_senstive_ignore:  # if the name is case_sensitive then check for exact value
                        if sheet not in self.ignore_sheets:
                            sheets.append(sheet)
                    elif sheet.lower() not in self.ignore_sheets:  # if checking is not case sensitive than checking value
                        sheets.append(sheet)                       # in lower as ignore list is also converted to lower
                else:
                    sheets.append(sheet)
        logger.debug("Sheets Gathered after filtering ignored sheets.")
        return sheets

    def get_change_logs(self, sheets, source, clone):
        # method returns a list of change log with all data in correct column position
        logs = []
        for sheet in sheets:
            source_sht = source.get_sheet_by_name(sheet)
            clone_sht = clone.get_sheet_by_name(sheet)
            headers, new_headers, removed_headers = self.get_headers(source_sht, clone_sht)
            for header in new_headers:
                # adding new headers in change log list with appropriate values
                logs.append([
                    datetime.now().strftime("%d-%m-%Y %H:%M"),
                    source_sht.title,
                    header,
                    1,
                    "-",
                    "(new header)"
                ])
            for header in removed_headers:
                # adding removed headers in change log list with appropriate values
                logs.append([
                    datetime.now().strftime("%d-%m-%Y %H:%M"),
                    source_sht.title,
                    header,
                    1,
                    "(removed header)",
                    "-"
                ])
            for header in headers:
                # looping through the headers which are to be considered in change log
                if header in new_headers or header in removed_headers:
                    # skipping header which are newly_add or removed so that change log will only have entry of header
                    # else it will be filled with each row data for the changed header, which might make file very huge
                    continue
                logs += self.find_changes(header, source_sht, clone_sht)
        logger.debug("Change logs gathered")
        return logs

    def update_change_logs(self, logs, clone):
        # Updates Chane Log sheet in log file
        sht = clone.get_sheet_by_name(self.log_sheet_name)
        for log in logs:  # looping through logs list which contains a lists of log with data inside in perfect order
            row = 1 + sht.max_row  # max row + 1 so the row is always added in the end of worksheet
            for j in range(1, len(log)+1):  # looping through index of log data, which is also used as column number
                sht.cell(row, j).value = log[j-1]
        logger.debug("Change logs sheet updated")

    def update_clone_file(self, source, clone, source_sheets):
        # updating clone file by looping through all sheets of source_file so that the only other data in clone file is
        # change log sheet, this updating contains each and every data including the headers and sheet which are ignored
        # for change log. If the sheet already present in clone file then we delete it and create a new one with exact
        # data from source
        for sheet in source_sheets:
            if sheet in clone.sheetnames:  # checking if the sheet exists in clone file, if it does than we first take its index than delete it
                clone_sht = clone.get_sheet_by_name(sheet)
                ind = clone.index(clone_sht)
                ignored_headers_data = self.get_ignored_data_from_clone(clone, sheet)  # ignored data from clone file
                clone.remove(clone_sht)
            else:  # if clone file doesn't has that sheet than make the index containing variable to "False"
                ind = "False"  # A string is taken instead of any empty value because the original index can contain 0
                               # which can cause conflict
            if ind != "False":  # if we have an index value than make a new sheet in that position
                clone.create_sheet(sheet, ind)
            else:  # else we will add the new sheet in second last position behind Change Log sheet
                clone.create_sheet(sheet, len(clone.sheetnames)-2)
            source_sht = source.get_sheet_by_name(sheet)
            clone_sht = clone.get_sheet_by_name(sheet)
            skipped = 0
            for row in range(1, source_sht.max_row+1):  # looping through every row from source
                for column in range(1, source_sht.max_column+1):  # looping through every column from source
                    if sheet in self.headers_column:
                        if column-1 in self.headers_column[sheet]:
                            if row == 1:  # if first row then write header
                                clone_sht.cell(row, column).value = self.headers_column[sheet][column-1]
                            elif row > 1:  # write data from the ignore data list, column-1 coz the header was index from 0 starting point
                                clone_sht.cell(row, column).value = ignored_headers_data[
                                    self.headers_column[sheet][column-1]][row-2]  # row-2 because here it starts from 2 and list starts from 0
                            continue
                    clone_sht.cell(row, column).value = source_sht.cell(row, column-skipped).value  #writing everything in clone
            for idx, rd in source_sht.row_dimensions.items():  # copying width and height of rows and columns
                clone_sht.row_dimensions[idx] = copy(rd)
            logger.debug(f"{sheet} sheet updated in clone file")
        logger.debug("All sheets updated in clone file")

    def find_changes(self, header, source_sht, clone_sht):
        # finding changes for a column(header) from source to clone
        changes = []
        # indexing header position of source and clone, incase if position of columns are shifted without changing data
        # than also our change log will consider it as change as the cell value has changed, so we are looking for
        # changes with header index so that less useful data is not stored in change log
        source_column = [cell.value for cell in source_sht[1]].index(header) + 1
        clone_column = [cell.value for cell in clone_sht[1]].index(header) + 1
        # getting max row number of source and clone and using the greater one to loop in all the files so that each
        # and every data is processed inside loop
        if source_sht.max_row > clone_sht.max_row:
            nrow = source_sht.max_row + 1
        else:
            nrow = clone_sht.max_row + 1
        for i in range(2, nrow):  # we have already processed headers so we have skipped it and started loop from 2
            source_value = source_sht.cell(i, source_column).value
            clone_value = clone_sht.cell(i, clone_column).value
            if clone_value != source_value:
                # if clone and source value didn't matched then creating a list with appropriate data in correct position
                changes.append([
                    datetime.now().strftime("%d-%m-%Y %H:%M"),
                    source_sht.title,
                    header,
                    i,
                    clone_value,
                    source_value
                ])
        return changes

    def get_headers(self, source_sht, clone_sht):
        # return a list of headers which are to be considered for change log, along with it return 2 others list with
        # new headers and removed headers so the change log contain directly entry of changes headers instead of whole
        # columns. Logic is same like get_sheets method
        source_headers = [cell.value for cell in source_sht[1]]
        clone_headers = [cell.value for cell in clone_sht[1]]
        self.headers_column[source_sht.title] = {}
        headers = []
        temp_headers = []
        new_headers = []
        removed_headers = []
        for header in source_headers:
            if not header:
                continue
            if self.ignore_headers:
                if self.case_senstive_ignore:
                    if header not in self.ignore_headers:
                        temp_headers.append(header)
                    else:
                        self.headers_column[source_sht.title][source_headers.index(header)] = header
                elif header.lower() not in self.ignore_headers:
                    temp_headers.append(header)
                else:
                    self.headers_column[source_sht.title][source_headers.index(header)] = header
            else:
                temp_headers.append(header)

        for header in temp_headers:
            if header not in clone_headers:
                new_headers.append(header)
            headers.append(header)

        for header in clone_headers:
            if not header:
                continue
            if self.ignore_headers:
                if self.case_senstive_ignore:
                    if header not in self.ignore_headers:
                        if header not in source_headers:
                            removed_headers.append(header)
                elif header.lower() not in self.ignore_headers:
                    if header not in source_headers:
                        removed_headers.append(header)
            else:
                if header not in source_headers:
                    removed_headers.append(header)
        logger.debug("Headers gathered after filtering ignored headers")
        return headers, new_headers, removed_headers

    def get_ignored_data_from_clone(self, clone, sheet):
        data = {}
        sht = clone.get_sheet_by_name(sheet)
        headers = [cell.value for cell in sht[1]]
        if not sheet in self.headers_column:
            return data
        for header in self.headers_column[sheet]:
            hd = self.headers_column[sheet][header]  # getting header from class list storing index of ignored headers
            data[hd] = []
            for i in range(2, sht.max_row+1):
                data[hd].append(sht.cell(i, headers.index(hd)+1).value)
        return data
