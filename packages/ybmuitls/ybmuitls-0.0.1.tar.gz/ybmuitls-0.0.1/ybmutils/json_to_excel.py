from openpyxl import Workbook


class ExportExcel(object):
    def __init__(self, path: str = "", filename: str = "example"):
        self.path = path
        self.filename = filename
        self.wb = Workbook()
        if self.wb.get_sheet_by_name('Sheet') is not None:
            self.wb.remove(self.wb.get_sheet_by_name('Sheet'))

    def add_sheet_by_json(self, sheet_name: str, data_list: list):
        ws = self.wb.create_sheet(sheet_name)
        keys = data_list[0].items()
        columns = []
        for column, key in enumerate(keys):
            ws.cell(row=1, column=column + 1, value=key[0])
            columns.append(key[0])

        for row, item in enumerate(data_list):
            for column, key in enumerate(columns):
                ws.cell(row=row + 2, column=column + 1, value=item[key])

    def add_sheet_by_json_with_key(self, sheet_name: str, data_list: list, head_list: list):
        ws = self.wb.create_sheet(sheet_name)
        for column, key in enumerate(head_list):
            ws.cell(row=1, column=column + 1, value=key)

        for row, item in enumerate(data_list):
            for column, key in enumerate(head_list):
                ws.cell(row=row + 2, column=column + 1, value=item[key])

    def add_sheet_by_json_with_headers(self, sheet_name: str, data_list: list, head_list: list):
        '''

        :param sheet_name:
        :param data_list: json list
        :param head_list:  list[tuple] (k,v)
        :return:
        '''
        ws = self.wb.create_sheet(sheet_name)
        k_list = []
        v_list = []
        for item in head_list:
            k_list.append(item[0])
            v_list.append(item[1])
        for index, key in enumerate(v_list):
            ws.cell(row=1, column=index + 1, value=key)
        for row, item in enumerate(data_list):
            for column, key in enumerate(k_list):
                row_key = item[key] if key in item else ""
                ws.cell(row=row + 2, column=column + 1, value=row_key)

    def save(self):
        self.wb.save(self.path + "/" + self.filename + ".xlsx")


def create_excel_by_json(path, filename, data_list):
    wb = Workbook()
    ws = wb.create_sheet(filename)
    keys = data_list[0].items()
    columns = []
    column = 0
    for key in keys:
        column += 1
        ws.cell(row=1, column=column, value=key[0])
        columns.append(key[0])

    for item in data_list:
        for key in columns:
            ws.cell(row=data_list.index(item) + 2, column=columns.index(key) + 1, value=item[key])

    wb.save(path + filename + '.xlsx')


def create_excel_by_json_with_head(path, filename, data_list, head_list):
    wb = Workbook()
    ws = wb.create_sheet(filename)
    for key in head_list:
        ws.cell(row=1, column=head_list.index(key) + 1, value=key)

    for item in data_list:
        for key in head_list:
            ws.cell(row=data_list.index(item) + 2, column=head_list.index(key) + 1, value=item[key])

    wb.save(path + filename + '.xlsx')

