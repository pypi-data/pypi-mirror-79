import pandas as pd
import openpyxl 
from pynvn.path.ppath import PathSteel
from tkinter import messagebox
class hdata:
    """hadling data excel file kid """
    def __init__(self,pathfile = None,
                sheetname = "AZB-10",
                rowindexstart = 41,
                columnmh = 3,
                columntotp = 18,
                valuecelltoskip = None,
                pathfile_to = None,
                sheetname_to = "AZB-10",
                rowindexstart_to = 41,
                columnmh_to = 3,
                columntotp_to = 18,
                valuecelltoskip_to = None,
                ):

                self.pathfile = pathfile
                self.sheetname = sheetname
                self.rowindexstart = rowindexstart
                self.columnmh = columnmh
                self.columntotp = columntotp
                self.valuecelltoskip = valuecelltoskip

                # start file to 
                self.pathfile_to = pathfile_to
                self.sheetname_to = sheetname_to
                self.rowindexstart_to = rowindexstart_to
                self.columnmh_to = columnmh_to
                self.columntotp_to = columntotp_to
                self.valuecelltoskip_to = valuecelltoskip_to

                # workbook object is created
                self.wb_obj = openpyxl.load_workbook(self.pathfile,
                                                            data_only=True) 

                self.wb_obj_to = openpyxl.load_workbook(self.pathfile_to,
                                                            data_only=True) 
                        
                # get sheet name file 
                self.sheet = self.wb_obj.get_sheet_by_name(self.sheetname)

                self.sheet_to = self.wb_obj_to.get_sheet_by_name(self.sheetname_to)

                # Max row
                self.mr = self.sheet.max_row
                #return arr value find 
                self.valuarr = self.reindexofvalue()
                self.valuarr.append (self.mr)

    def hldatakid(self):
        """hadling data excel file kid """
        # sum for caculation 
        self.calsumvaluecolumntotp(self.valuarr)
        try:
            self.wb_obj.save(self.pathfile)
        except:
            messagebox.showerror("error","this file is openning, close it")

    def hldataparent(self):
        """hadling data excel file parent """
        revalu = self.returnvluenotnone()
        for inrowvalue in revalu:
            retuar = self.returnvaluebyvalueothesheet(inrowvalue)
            
            self.sheet_to.cell(row=retuar[0], column=self.columntotp_to).value = retuar[1]

        self.wb_obj_to.save(self.pathfile_to)

    def returnvaluebyvalueothesheet(self,inrowvalue):
        """ return value by value the sheet """

        for rowindex in range (self.rowindexstart,self.mr):
            varow = self.sheet.cell(row=rowindex, 
                                    column=self.columnmh).value

            valrowget = self.sheet.cell(row=rowindex, 
                                    column=self.columntotp).value
            if varow==inrowvalue:
                return [rowindex,valrowget]
                break
        else:
            messagebox.showinfo("Title", "Not found value {}").format(rowindex)
    
    def reindexofvalue(self):
        """ return index of value"""
        i = 0 
        a = True
        valuearr = []
        valuearr1 = []
        cli = self.rowindexstart 
        while cli < self.mr:
            valcell = self.sheet.cell(row=cli, column=self.columnmh).value
            if valcell == self.valuecelltoskip:
                for ci in range (cli,self.mr):
                    vaulenone = self.sheet.cell(row=ci, 
                                                column=self.columnmh).value
                    if  vaulenone == self.valuecelltoskip:
                        i += 1
                    else:
                        valuearr.append (cli - 1)
                        cli = cli + i
                        a = False
                        i = 0
                        break
            if a == True:             
                cli += 1
            a = True
        return valuearr

    def returnvluenotnone(self):
        """ return value not none """
        i = 0 
        a = True
        valuearr = []
        cli = self.rowindexstart 
        while cli < self.mr:
            valcell = self.sheet.cell(row=cli, column=self.columnmh).value
            if valcell == self.valuecelltoskip:
                for ci in range (cli,self.mr):
                    vaulenone = self.sheet.cell(row=ci, 
                                                column=self.columnmh).value
                    if  vaulenone == self.valuecelltoskip:
                        i += 1
                    else:
                        valcell = self.sheet.cell(row=cli - 1, column=self.columnmh).value
                        valuearr.append (valcell)
                        cli = cli + i
                        a = False
                        i = 0
                        break
            if a == True:             
                cli += 1
            a = True
        return valuearr

    def calsumvaluecolumntotp (self,listvalue):
        """ caclulate value column """
        valnotnone = self.findcellnotnone()
        sum = 0
        valuearr = list(self.func(listvalue))
        for elelist in valuearr:

            relistn = self.remiveinter(list(range(elelist[0] + 1,elelist[1])),
                                    valnotnone)
            relistn.sort()
            
            for idex in relistn:
                value = self.sheet.cell(row=idex, 
                                        column=self.columntotp).value
                valuem = value if value is not None else 0 
                sum += valuem
                self.sheet.cell(row=min(relistn) - 1, 
                                column=self.columntotp).value = sum
            sum = 0

    def func(self,alist):

        return zip(alist, alist[1:])

    def intersertionlist(self,lst1, lst2):
        lst3 = [value for value in lst1 if value in lst2] 
        return lst3 

    def findcellnotnone(self):
        """ find cell not none value """
        valnotnone = []
        cli = self.rowindexstart 
        while cli < self.mr:
            valcell = self.sheet.cell(row=cli, 
                                    column=self.columnmh).value
            if valcell != None:
                valnotnone.append (cli)
            cli += 1
        return valnotnone

    def remiveinter (self,listmn,listmned):
        """ return value intersection list """
        interlist  = self.intersertionlist(listmned,
                                            listmn)
        return list(set(listmn)- set(interlist))
