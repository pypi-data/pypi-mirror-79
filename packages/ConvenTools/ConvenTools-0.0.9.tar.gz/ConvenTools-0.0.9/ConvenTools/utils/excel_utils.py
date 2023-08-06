import xlrd
#获取excel数据返回字典对象
import openpyxl
from openpyxl import Workbook



def readExcel_to_json(filename,sheetsIndex):
    excel = xlrd.open_workbook(filename)
    sheet1 = excel.sheets()[sheetsIndex]
    # 找到有几列几列
    nrows = sheet1.nrows  # 行数
    ncols = sheet1.ncols  # 列数
    totalArray = []
    title = []
    # 标题
    for i in range(0, ncols):
        title.append(sheet1.cell(0, i).value);
    # 数据
    for rowindex in range(1, nrows):
        dic = {}
        for colindex in range(0, ncols):
            s = sheet1.cell(rowindex, colindex).value
            dic[title[colindex]] = s
        totalArray.append(dic);
    return totalArray




#获取excel数据返回字典对象
def readExcel_By_sheetName_to_json(filename,sheeetNam):
    excel = xlrd.open_workbook(filename,"rb")
    sheet1 = excel.sheet_by_name(sheeetNam)
    # 找到有几列几列
    nrows = sheet1.nrows  # 行数
    ncols = sheet1.ncols  # 列数
    totalArray = []
    title = []
    # 标题
    for i in range(0, ncols):
        title.append(sheet1.cell(0, i).value);
    # 数据
    for rowindex in range(1, nrows):
        dic = {}
        for colindex in range(0, ncols):
            s = sheet1.cell(rowindex, colindex).value
            dic[title[colindex]] = s
        totalArray.append(dic);
    return totalArray


#把json写入excel
def writeExcel_to_json(filename,jsonData,sheetName):
    """
    创建文件，并将数据写入到默认的sheet里面，并重命名sheet
    :param filename:
    :param jsonData:
    :param sheetName:
    :return:
    """
    wb =  Workbook()
    sheet =  wb.active
    sheet.title = sheetName
    # 找到有几列
    ncols = len(jsonData[0])
    totalArray = []
    title = []
    # 标题
    titles = list(jsonData[0].keys())
    for i in range(0, ncols):
        sheet.cell(0 + 1, i + 1, titles[i])
    # 数据
    for rowindex in range(1, len(jsonData)+1):
        for colindex in range(0, len(titles)):
            try:
                sheet.cell(rowindex + 1, colindex + 1, jsonData[rowindex - 1][titles[colindex]])
            except Exception as e:
                print("当前错误数据：")
                print(jsonData[rowindex - 1])
                print("类型为：")
                print(e)
    wb.save(filename)  # 保存工作簿


#把json写入到指定的excel sheet(叠加保存之前数据)
def write_And_Copy_Excel_to_json(filename,jsonData,sheetName):
    """
    把json写入到指定的excel sheet(叠加保存之前数据)  追加sheet
    :param filename:
    :param jsonData:
    :param sheetName:
    :return:
    """
    # filename = filename.decode('utf-8')
    wb = openpyxl.load_workbook(filename)
    sheet =  wb.create_sheet()
    sheet.title = sheetName
    # 找到有几列
    ncols =  len(jsonData[0])
    totalArray = []
    title = []
    # 标题
    titles =  list(jsonData[0].keys())
    for i in range(0, ncols):
        sheet.cell(0+1, i+1, titles[i])
    # 数据
    for rowindex in range(1, len(jsonData)+1):
        for colindex in range(0, len(titles)):
            try:
                sheet.cell(rowindex+1, colindex+1, jsonData[rowindex-1][titles[colindex]])
            except Exception as e:
                print("当前错误数据：")
                print(jsonData[rowindex-1])
                print("类型为：")
                print(e)
    wb.save(filename)  # 保存工作簿


from ConvenTools.utils import file_utils


#写入二维数组到指定文件，指定sheet
def writeExcel_to_matrix(filename,matrixData,sheetName):
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
        # 如果文件不存在，就创建文件
    file_utils.mknod(filename)
    wb = openpyxl.load_workbook(filename)
    sheet = wb.create_sheet()
    sheet.title = sheetName
    for rowIndex in range(len(matrixData)):
        for colindex in range(len(matrixData[rowIndex])):
            try:
                sheet.cell(rowIndex + 1, colindex + 1, matrixData[rowIndex][colindex])
            except Exception as e:
                print("出错了")
    wb.save(filename)



def creatEmptyFile(fileName):
    """
    创建空的excel文件
    :param fileName:
    :return:
    """
    wb = Workbook()
    wb.save(fileName)
