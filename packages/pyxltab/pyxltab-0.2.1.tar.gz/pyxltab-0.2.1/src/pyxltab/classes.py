"""
Extensions to `openpyxl` classes allowing for the cells corresponding to table columns
to be returned.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from openpyxl.cell.cell import Cell as openpyxl_Cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook as openpyxl_Workbook
from openpyxl.worksheet.table import Table as openpyxl_Table
from openpyxl.worksheet.table import TableColumn as openpyxl_TableColumn
from openpyxl.worksheet.worksheet import Worksheet as openpyxl_Worksheet


class ChildrenMapping(Mapping):
    """
    A `dict`-like class that indexes into one of its attributes designated as its
    `children`. The attribute must be a `dict` itself. Allows returning children like
    `ChildrenMapping["index"]`. Can iterate over keys and values of children like `for
    (key, value) in ChildrenMapping:...`, or just the values like `for value in
    ChildrenMapping.values():`. The length of a `ChildrenMapping` is the length of its
    `children` dictionary.
    """

    def __init__(
        self,
        children: Dict[str, Any],
    ):
        self._children = children

    def __getitem__(self, key: str):
        return self._children[key]

    def __iter__(self):
        return iter(self._children)

    def __len__(self):
        return len(self._children)


class Book(ChildrenMapping):
    """
    Extends the `openpyxl.workbook.workbook.Workbook` class from `openpyxl`.
    """

    def __init__(self, openpyxl_book: openpyxl_Workbook):
        self.openpyxl_book = openpyxl_book
        self.sheets = {
            openpyxl_sheet.title: Sheet(openpyxl_sheet, parent=self)
            for openpyxl_sheet in openpyxl_book.worksheets
        }
        super().__init__(children=self.sheets)


class Sheet(ChildrenMapping):
    """
    Extends the `openpyxl.worksheet.worksheet.Worksheet` class from `openpyxl`.
    """

    def __init__(self, openpyxl_sheet: openpyxl_Worksheet, parent: Book):
        self.parent = parent
        self.openpyxl_sheet = openpyxl_sheet
        self.tables = {
            openpyxl_table.name: Table(openpyxl_table, parent=self)
            for openpyxl_table in openpyxl_sheet.tables.values()
        }
        super().__init__(children=self.tables)


class Table(ChildrenMapping):
    """
    Extends the `openpyxl.worksheet.table.Table` class from `openpyxl`.
    """

    def __init__(self, openpyxl_table: openpyxl_Table, parent: Sheet):
        self.parent = parent
        self.openpyxl_table = openpyxl_table
        (self.first_col, self.first_row, _, self.last_row) = range_boundaries(
            openpyxl_table.ref
        )
        self.columns: Dict[str, Column] = {}

        for openpyxl_column in openpyxl_table.tableColumns:
            self.columns[openpyxl_column.name] = Column(openpyxl_column, parent=self)

        super().__init__(children=self.columns)

    def get_cells(self) -> List[ColumnCells]:
        """
        Get the cells in the table as a list of `ColumnCells` objects.
        """

        table_cells = [column.get_cells() for column in self.values()]
        return table_cells


class Column(Sequence):
    """
    Extends the `openpyxl.worksheet.table.TableColumn` class from `openpyxl`.
    """

    def __init__(self, openpyxl_column: openpyxl_TableColumn, parent: Table):
        self.parent = parent
        self.openpyxl_column = openpyxl_column

        self.col_num = parent.first_col + len(parent.columns)

        self.column_cells: Optional[ColumnCells] = None
        self.header: Optional[openpyxl_Cell] = None
        self.cells: Optional[List[openpyxl_Cell]] = None
        self.total: Optional[openpyxl_Cell] = None

    def __getitem__(self, index):
        """
        Subscript into the `cells` attribute, getting the cells first if necessary.
        """

        if self.cells is None:
            self.get_cells()
        return self.column_cells[index]

    def __len__(self):
        """
        Get the length of the `cells` attribute, getting the cells first if necessary.
        """

        if self.cells is None:
            self.get_cells()
        return len(self.column_cells)

    def get_cells(self) -> ColumnCells:
        """
        Get cells in this column as a `ColumnCells` object.
        """

        table = self.parent
        header_row_count = table.openpyxl_table.headerRowCount
        totals_row_count = table.openpyxl_table.totalsRowCount

        sheet = table.parent

        if header_row_count == 1:
            header = sheet.openpyxl_sheet.cell(row=table.first_row, column=self.col_num)
        elif header_row_count is None or header_row_count == 0:
            header = None
            header_row_count = 0

        if totals_row_count == 1:
            total = sheet.openpyxl_sheet.cell(row=table.last_row, column=self.col_num)
        elif totals_row_count is None or totals_row_count == 0:
            total = None
            totals_row_count = 0

        between = next(
            sheet.openpyxl_sheet.iter_cols(
                min_col=self.col_num,
                max_col=self.col_num,
                min_row=table.first_row + header_row_count,
                max_row=table.last_row - totals_row_count,
            )
        )

        self.column_cells = ColumnCells(header, between, total)
        self.header = self.column_cells.header
        self.cells = self.column_cells.between
        self.total = self.column_cells.total

        return self.column_cells


@dataclass
class ColumnCells(Sequence):
    """
    Contains a column of cells from an Excel table.
    """

    header: Optional[openpyxl_Cell]
    between: List[openpyxl_Cell]
    total: Optional[openpyxl_Cell]

    def __getitem__(self, index):
        """
        Subscript into the `between` attribute.
        """

        return self.between[index]

    def __len__(self):
        """
        Get the length of the `between` attribute.
        """

        return len(self.between)
