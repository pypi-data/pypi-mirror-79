"""
This module is dedicated to the conversion from outside format to internal json format.
Outside formats may be : a workbook (excel), another json file, a database etc...
Structure and specifications of internal json format are defined in this module. Internal
json format can take two main forms : one to adress input informations and a second one
for output communications.
"""

import os
import tempfile
import xmltodict
import shutil

import pandas as pd
import numpy as np

from zipfile import ZipFile
from openpyxl import load_workbook
from collections import OrderedDict
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font

try:
    from . import io_bdd
except Exception:
    import io_bdd
try:
    from . import su_trace
except Exception:
    import su_trace


def load_mfa_problem_from_excel(
    input_file: str,
    create_empty_ter=False  # if True only products and sectors worksheets will be extracted
):
    '''Main convertor routine. Call dedicated routine depending on input type
    - input_file : string with input file name (with extension and path)
    '''
    input_categories = [
        'param', 'dim_products', 'dim_sectors', 'data',
        'min_max', 'constraints', 'ter_base'
    ]

    df_ifile = {}
    tab_list = []
    df_ifile = pd.read_excel(input_file, None)
    real_tablist = [*df_ifile]
    consistant_tab_list = ['prod', 'sect']
    if not create_empty_ter:
        consistant_tab_list += [
            'param', 'flux', 'data', 'minmax',
            'constr', 'proxy', 'pflow', 'psect'
        ]
    # keeping only consistent sheets
    for real_tab in real_tablist:
        consistent_sheet = consistantSheetName(real_tab)
        if consistent_sheet in consistant_tab_list:
            # real_tab is consistent
            tab_list.append(real_tab)
            if (consistent_sheet in ['flux']):
                # dtype has to be str to keep geographic ids (of Communes for exemple) in
                # string format => reload flux worksheet with the correct dtype
                del df_ifile[real_tab]
                df_ifile[real_tab] = pd.read_excel(input_file, sheet_name=real_tab, dtype=str)
    su_trace.logger.debug('Names of excel sheet :' + str(tab_list))
    mfa_problem_input = xl_convert_tablist(df_ifile, tab_list)
    # add eventual missing entries
    for input_category in input_categories:
        if input_category not in mfa_problem_input.keys():
            mfa_problem_input[input_category] = []
    return mfa_problem_input


def consistantSheetName(
    prop_sheet: str
):
    '''
    Test if the prop_sheet is consistent with the allowed sheet list.
    - Result is empty string if the tested sheet is not consistant.
    - Result is the dictionary key corresponding of the allowed list found.
    Note 1 : if the prop_sheet input is empty ('') the result is a list of
    allowed sheet name as a string
    Note 2 : a particular case is taken into account for proxy input file which
    usualy has 3 proxy sheets (and one of them with 'sector' keyword in its name)
    '''
    dictofsheetsnames = {
        'param': ['param'], 'prod': ['produit', 'product'],
        'sect': ['secteur', 'sector'], 'geo': ['geo', 'liste'],
        'flux': ['ter', 'flux'],
        'data': ['data', 'donn'], 'minmax': ['min max', 'min_max'],
        'constr': ['contrainte', 'constraint'],
        'proxy': ['proxi', 'proxy'], 'pflow': ['flow'], 'psect': []
    }
    prop_sheet = prop_sheet.lower()
    list_allowed = ''
    if prop_sheet != '':
        particular_case = False
        for allow_sheet in dictofsheetsnames['proxy']:
            if allow_sheet in prop_sheet:
                particular_case = True
        if particular_case:
            # We are in the particular case (cf Note 2)
            for allow_sheet in dictofsheetsnames['pflow']:
                if allow_sheet in prop_sheet:
                    return 'pflow'
            for allow_sheet in dictofsheetsnames['sect']:
                if allow_sheet in prop_sheet:
                    return 'psect'
            return 'proxy'
        else:
            for dict_key in dictofsheetsnames.keys():
                for allow_sheet in dictofsheetsnames[dict_key]:
                    if allow_sheet in prop_sheet:
                        return dict_key
    else:
        for dict_key in dictofsheetsnames.keys():
            if len(dictofsheetsnames[dict_key]) != 0:
                list_allowed += ', '.join(dictofsheetsnames[dict_key])
    return list_allowed


def xl_import_tab(
    df_fi: dict,
    stab: str,
    def_val: list,
    js_tab: str,
    mfa_problem_input: dict
):
    """Import informations from workbook tab called stab if it exists
    - df_fi : dataframe with all sheets of the input file
    - stab : name of the workbook sheet to work on
    - def_val : dictionnary of default values (default columns values of excel sheet)
    - js_tab : name of the main JSon dictionnary key for this entry
    - mfa_problem_input : dictionnary with informations to convert in JSon format
    """
    my_json = []
    try:
        df_prod = df_fi[stab]
        # Check if we need to add empty columns
        if len(list(df_prod)) < len(def_val):
            starti = len(list(df_prod))
            for i in range(starti, len(def_val)):
                cname = "Col_" + str(i)
                df_prod[cname] = def_val[i]
        # Fill dataframe nan with default value
        for i, col in enumerate(list(df_prod)):  # iterable on columns names
            vdef = ''
            if i < len(def_val):
                vdef = def_val[i]
            if vdef is not None:
                df_prod[col] = df_prod[col].fillna(value=vdef)
                if type(vdef) is str:
                    df_prod[col] = df_prod[col].astype(str)
                if type(vdef) is int:
                    df_prod[col] = df_prod[col].astype(int)
            else:
                df_prod[col] = df_prod[col].replace({np.nan: None})
        # Extract values (nparray) from dataframe and convert them in a list format
        my_json = df_prod.values.tolist()
    except Exception:
        su_trace.logger.error(f'Exception in xl_import_tab for {js_tab} entry')
    mfa_problem_input[js_tab] = my_json


def input_to_json(
    input_type,
    input_file,
    sess_act,
    mod_name
):
    '''Main convertor routine. Call dedicated routine depending on input type
    - input_type : type of the input (0 : xls/xlsx/csv, 1: database, 2: JSon)
    - input_file : string with input file name (with extension and path)
    - xltab_list : list of main entries awaited
    - jstab_list : list of entries needed in JSon file
'''
    # Names of tab sheets as they should be after naming normalisation
    jstab_list = [
        'param', 'dim_products', 'dim_sectors', 'data', 'min_max',
        'constraints', 'ter_base'
    ]

    mfa_problem_input = {}
    if input_type == 0:  # excel input
        # import input_file only once
        df_ifile = pd.read_excel(input_file, None)
        myxl_tablist = [*df_ifile]  # myxl_tablist = list(df_ifile.keys()) is slower
        su_trace.logger.debug('Names of excel sheet :' + str(myxl_tablist))
        mfa_problem_input = xl_convert_tablist(df_ifile, myxl_tablist)
        # add eventual missing entries
        for tab in jstab_list:
            if tab not in mfa_problem_input.keys():
                mfa_problem_input[tab] = []
    elif input_type == 2:  # input from data base
        mfa_problem_input['param'] = io_bdd.read_inputs(io_bdd.Param, sess_act, mod_name, '', '')
        mfa_problem_input['dim_products'] = io_bdd.read_inputs(io_bdd.Product, sess_act, mod_name, '', '')
        mfa_problem_input['dim_sectors'] = io_bdd.read_inputs(io_bdd.Sector, sess_act, mod_name, '', '')
        mfa_problem_input['ter_base'] = io_bdd.read_inputs(io_bdd.Flux, sess_act, mod_name, '', '')
        mfa_problem_input['data'] = io_bdd.read_inputs(io_bdd.Data, sess_act, mod_name, '', '')
        mfa_problem_input['min_max'] = io_bdd.read_inputs(io_bdd.MinMax, sess_act, mod_name, '', '')
        mfa_problem_input['constraints'] = io_bdd.read_inputs(io_bdd.Constraint, sess_act, mod_name, '', '')
    else:
        pass
    # add eventual missing entries
    for tab in jstab_list:
        if tab not in mfa_problem_input.keys():
            mfa_problem_input[tab] = []
    return mfa_problem_input


def xl_get_sheet_details(
    file_path,
    only_sheets=True
):
    '''
    Finded at : https://stackoverflow.com/questions/17977540/pandas-looking-up-the-list-of-sheets-in-an-excel-file
    Speedest way to get informations from an excel file without the nead to open it
    Benchmarking: (On a 6mb xlsx file with 4 sheets)
    Pandas, xlrd: 12 seconds
    openpyxl: 24 seconds
    Proposed method: 0.4 seconds
    Notes (modifications I made):
    - use tempfile.mkdtemp instead of settings.MEDIA_ROOT
    - routine adapted to extract only sheets names (when entry only_sheets=True)
    Requirements :
    - must install xmltodict and add 'import xmltodict'
    - must add 'import tempfile'
    - must add 'import shutil'
    - must add 'from zipfile import ZipFile'
    '''
    sheets = []
    # Make a temporary directory to work in
    directory_to_extract_to = tempfile.mkdtemp()
    # Extract the xlsx file as it is just a zip file
    zip_ref = ZipFile(file_path, 'r')
    zip_ref.extractall(directory_to_extract_to)
    zip_ref.close()

    # Open the workbook.xml which is very light and only has meta data, get sheets from it
    path_to_workbook = os.path.join(directory_to_extract_to, 'xl', 'workbook.xml')
    with open(path_to_workbook, 'r', encoding='utf8') as f:
        xml = f.read()
        dictionary = xmltodict.parse(xml, encoding='utf-8')
        for sheet in dictionary['workbook']['sheets']['sheet']:
            sheet_details = {
                'id': sheet['@sheetId'],
                'name': sheet['@name']
            }
            if only_sheets:
                sheets.append(sheet['@name'])
            else:
                sheets.append(sheet_details)

    # Delete the extracted files directory
    shutil.rmtree(directory_to_extract_to)
    return sheets


def xl_import_param(
    df_fi: dict,
    stab: str,
    mfa_problem_input: dict
):
    """Import information from workbook tab called "param" if it exists
    - df_fi : dataframe with all sheets of the input file
    - stab : name of the workbook tabulation to work on
    - mfa_problem_input : dictionnary with informations to convert in JSon format
    """
    my_json = []
    try:
        df_prod = df_fi[stab]
        # Fill dataframe nan with default value
        df_prod = df_prod.fillna(value='').replace({np.nan: None})
        # Extract values (nparray) from dataframe and convert them in a list format
        my_json = df_prod.values.tolist()
    except Exception:
        su_trace.logger.error('Exception in xl_import_param')
    # mfa_problem_input['param'] = {'max': le_max, 'tol': tol}
    mfa_problem_input['param'] = {}
    for row in my_json:  # iterable on columns names
        mfa_problem_input['param'][row[0]] = row[1]


def xl_import_terbase(
    df_fi: dict,
    stab: str,
    mfa_problem_input: dict
):
    """Import informations from workbook tab called "ter_base" if it exists
    - df_fi : dataframe with all sheets of the input file
    - stab : name of the workbook tabulation to work on
    - mfa_problem_input : dictionnary with informations to convert in JSon format
    """
    my_json = {}
    try:
        li_tmp = [li[1] for li in mfa_problem_input['dim_products']]  # list of products (nb of rows)
        # list of UNIQUE products ordered by apparition order
        li_prod = list(OrderedDict.fromkeys(li_tmp))
        nb_prod = len(li_prod)  # number of UNIQUE products
        # list of UNIQUE sectors ordered by apparition order
        li_tmp = [li[1] for li in mfa_problem_input['dim_sectors']]
        li_sect = list(OrderedDict.fromkeys(li_tmp))
        # loading dataframe from excel file
        df_prod = df_fi[stab]
        li_col = df_prod.iloc[0, 2:].values.tolist()  # Get columns list
        li_row = df_prod.iloc[1:, 1].values.tolist()  # Get row indexes names
        df_prod = df_prod.iloc[1:, 2:]  # Values
        df_prod.index = li_row
        df_prod.columns = li_col
        df_prod = df_prod.loc[~df_prod.index.isnull()]  # remove all lines with empty index
        # We need to check if the nb_prod row still contains column names
        # (because in some excel file this line hasn't an empty index)
        if df_prod.iloc[nb_prod, 0] in li_sect:
            df_prod = df_prod.drop(df_prod.index[nb_prod])
        li_xlsect = list(df_prod)  # list of sectors (columns) in excel file
        # Check if the column order is consistent (well ordered)
        if li_xlsect != li_sect:
            # We need to rearange columns
            try:
                df_prod = pd_sorted_col(df_prod, li_sect)
            except Exception:
                dupes = [x for n, x in enumerate(df_prod.keys()) if x in df_prod.keys()[:n]]
                if len(dupes) > 0:
                    su_trace.logger.error(f'ERROR: Duplicate sectors in ter1. {dupes}')
                su_trace.logger.info('Sectors defined in dim that are not in ter1:')
                for k in li_sect:
                    if k not in df_prod.keys():
                        su_trace.logger.error(f'ERROR {k}')
                su_trace.logger.info('Sectors in ter1 that are not defined in dim:')
                for k in df_prod.keys():
                    if k not in li_sect:
                        su_trace.logger.error(f'ERROR {k}')
                df_prod = pd_sorted_col(df_prod, li_sect)
        # Extract "supply" part of the dataframe
        df_tmp = df_prod[0:nb_prod]
        li_xlprod = list(df_tmp.index.values)
        if li_xlprod != li_prod:
            # We need to rearange rows
            dft = df_tmp.transpose()
            try:
                dft = pd_sorted_col(dft, li_prod)
            except Exception:
                dupes = [x for n, x in enumerate(dft.keys()) if x in dft.keys()[:n]]
                if len(dupes) > 0:
                    su_trace.logger.error(f'ERROR: Duplicate products in ter1. {dupes}')
                su_trace.logger.error('Products defined in dim that are not in ter1:')
                for k in li_prod:
                    if k not in dft.keys():
                        su_trace.logger.error(f'ERROR: {k}')
                su_trace.logger.error('Prodcuts in ter1 that are not defined in dim:')
                for k in dft.keys():
                    if k not in li_prod:
                        su_trace.logger.error(f'ERROR: {k}')
                dft = pd_sorted_col(dft, li_prod)
            df_tmp = dft.transpose()
        # Extract values (nparray) from dataframe and convert them in a list format
        li_tmp = df_tmp.values.tolist()
        # Replace Nan by None
        li_clean = [[None if val != val else int(val) for val in lign] for lign in li_tmp]
        my_json['supply'] = li_clean
        # Extract "use" part of the dataframe
        ini = nb_prod
        df_tmp = df_prod[ini:ini+nb_prod]
        li_xlprod = list(df_tmp.index.values)
        if li_xlprod != li_prod:
            dft = df_tmp.transpose()
            dft = pd_sorted_col(dft, li_prod)
            df_tmp = dft.transpose()
        li_tmp = df_tmp.values.tolist()
        li_clean = [[None if val != val else int(val) for val in lign] for lign in li_tmp]
        my_json['use'] = li_clean
    except Exception as expt:
        su_trace.logger.error('Exception in xl_import_terbase: ' + str(expt))
    mfa_problem_input['ter_base'] = my_json


def pd_sorted_col(
    dft,
    lico
):
    """Sort columns order of a dataframe in function of a column list"""
    li_df = list(dft)
    if li_df != lico:
        dftm = pd.DataFrame(columns=lico)
        for col in lico:
            dftm[col] = dft[col]
    return dftm


def xl_convert_tablist(
    df_file: str,
    tab_list: list
):
    """ Convert each tab of a workbook in mfa_problem_input dictionnary entry
    - df_file : dataframe with all sheets of the input file
    - tab_list : input file worksheet list
    """
    mfa_problem_input = {}
    for tab in tab_list:
        consistant_tab = consistantSheetName(tab)
        if consistant_tab == 'param':
            xl_import_param(df_file, tab, mfa_problem_input)
        elif consistant_tab == 'prod':
            jstab = 'dim_products'
            if 'dim_products' not in mfa_problem_input:
                # List of columns in "pommes-poires.xlsx" exemple: level, Element,
                # Bilan matière ?, transport interreg,
                # poids consolidation (1 par défaut), table consolidation,
                # Sankey ?, Couleur
                valdef = [1, '', False, False, None, None, False, '']
                xl_import_tab(df_file, tab, valdef, jstab, mfa_problem_input)
        elif consistant_tab == 'sect':
            jstab = 'dim_sectors'
            if 'dim_sectors' not in mfa_problem_input:
                # List of columns in "pommes-poires.xlsx" exemple: level, Element,
                # Bilan matière ?, transport interreg,
                # poids consolidation (1 par défaut), table consolidation,
                # Sankey ?, Couleur
                valdef = [1, '', False, False, None, None, False, '']
                xl_import_tab(df_file, tab, valdef, jstab, mfa_problem_input)
        elif consistant_tab == 'data':
            jstab = 'data'
            if 'data' not in mfa_problem_input:
                # List of columns in "pommes-poires.xlsx" exemple: période, Région,
                # Table, Origine, Destination, Quantité, Incertitude (%),
                # Contrainte 2 sigmas
                valdef = ['', '', '', '', '', None, None, None, '', '', '', '']
                xl_import_tab(df_file, tab, valdef, jstab, mfa_problem_input)
                # valdef = ['', '', '', '', '', None, None, '', None, '', None, '']
        elif consistant_tab == 'minmax':
            jstab = 'min_max'
            if 'min_max' not in mfa_problem_input:
                # List of columns in "pommes-poires.xlsx" exemple: Période, Région,
                # Table, Origine, Destination, min, max
                valdef = ['', '', '', '', '', None, None, '', '', '', '', '']
                xl_import_tab(df_file, tab, valdef, jstab, mfa_problem_input)
                # valdef = ['', '', '', '', '', None, None, '', '', '', None, '']
        elif consistant_tab == 'constr':
            jstab = 'constraints'
            if 'constraints' not in mfa_problem_input:
                # List of columns in "pommes-poires.xlsx" exemple: id, Période, Région,
                # Ressources/Emplois, Origine, Destination, eq = 0, eq <= 0, eq >= 0
                valdef = [None, '', '', '', '', '', None, None, None]
                xl_import_tab(df_file, tab, valdef, jstab, mfa_problem_input)
        elif consistant_tab == 'flux':
            if 'ter_base' not in mfa_problem_input:
                xl_import_terbase(df_file, tab, mfa_problem_input)
        else:
            pass
    return mfa_problem_input


def write_mfa_problem_output_to_excel(
    output_file_name: str,
    mfa_problem_input: dict,
    mfa_problem_output: dict
):
    with pd.ExcelWriter(output_file_name, engine='openpyxl', mode='a') as writer:
        for tab_name, tab_content in mfa_problem_output.items():
            sheet_content = tab_content
            # We don't want to write all sub territories TER results
            li_tmp = [co[1] for co in mfa_problem_input['dim_sectors']]  # sectors list
            sectors_names = list(OrderedDict.fromkeys(li_tmp))
            write_tab = True
            if 'ter' in tab_name:
                excluded_list = [sect for sect in sectors_names if sect in tab_name]
                if len(excluded_list) != 0:
                    write_tab = False
            if write_tab:
                if ((('ter' in tab_name) or ('flux' in tab_name)) and (len(tab_content) > 0)):
                    su = np.array(tab_content['supply'])
                    nb_rows = su.shape[0]
                    use = np.array(tab_content['use'])
                    df = pd.DataFrame(su)
                    df.to_excel(
                        writer, sheet_name=tab_name, index=False, header=False,
                        startrow=1, startcol=1
                    )
                    df = pd.DataFrame(use)
                    df.to_excel(
                        writer, sheet_name=tab_name, index=False, header=False,
                        startrow=nb_rows+3, startcol=1
                    )
                    format_excel(writer, tab_name, mfa_problem_input)
                else:
                    df = pd.DataFrame(sheet_content)
                    df.to_excel(writer, sheet_name=tab_name, index=False, header=False)


def format_excel(
    excel_writer: pd.ExcelWriter,
    tab_name: str,
    mfa_problem_input: dict
):
    li_tmp = [co[1] for co in mfa_problem_input['dim_products']]  # products list
    products_names = list(OrderedDict.fromkeys(li_tmp))
    li_tmp = [co[1] for co in mfa_problem_input['dim_sectors']]  # sectors list
    sectors_names = list(OrderedDict.fromkeys(li_tmp))
    col_name = ''
    n = len(sectors_names) + 2
    while n > 0:
        n, r = divmod(n - 1, 26)
        col_name = chr(r + ord('A')) + col_name

    mysheet = excel_writer.sheets[tab_name]

    greybg = PatternFill(bgColor='C0C0C0')
    greyft = Font(color='C0C0C0')
    rule = CellIsRule(operator='equal', formula=['0'], font=greyft, fill=greybg)
    srange = 'C3:' + col_name + str(2+len(products_names))
    mysheet.conditional_formatting.add(srange, rule)
    srange = 'C' + str(6+len(products_names)) + ':' + col_name + str(5+2*len(products_names))
    mysheet.conditional_formatting.add(srange, rule)


def excel_proxy_to_json(
    input_file: str,
    upper_level_name: str
):
    # _____INIT_____
    df_input_file = {}
    dict_tab_list = {}
    proxy_input = {}
    df_input_file = pd.read_excel(input_file, None)
    real_tablist = [*df_input_file]
    # _____CONFIG_____
    consistant_tab_list = ['geo', 'proxy', 'pflow', 'psect', 'data']
    # keeping only consistent sheets
    for real_tab in real_tablist:
        consistent_sheet = consistantSheetName(real_tab)
        if consistent_sheet in consistant_tab_list:
            # real_tab is consistent
            dict_tab_list[consistent_sheet] = real_tab
    # _____FILL_____
    df_input_data = df_input_file[dict_tab_list['data']]
    mask = df_input_data['region'] == upper_level_name
    input_data = np.array(df_input_data[mask].values)
    proxy_input['data'] = input_data
    proxy_input['proxis_flows'] = np.array(df_input_file[dict_tab_list['pflow']].values)
    proxy_input['proxis_sectors'] = np.array(df_input_file[dict_tab_list['psect']].values)
    regions = np.array(df_input_file[dict_tab_list['geo']].values)
    proxy_input['regions'] = [r[0] for r in regions]
    proxy_input['main_reg'] = upper_level_name
    proxy_input['years'] = [input_data[0, 0]]
    proxis = np.array(df_input_file[dict_tab_list['proxy']].values)
    proxy_input['proxis'] = proxis[:, [0, 1, 2]]
    # building data_ps (used for proxis sectors)
    data_ps = np.append(input_data[:, 0:1].astype(str), input_data[:, 1:], axis=1)
    ps = np.array(()).reshape((0, 2))
    for r in input_data:
        if r[2] in ['R', 'r', 'S', 's']:
            ps = np.append(ps, np.array([[r[4], r[3]]]), axis=0)
        else:
            ps = np.append(ps, np.array([[r[3], r[4]]]), axis=0)
    proxy_input['data_ps'] = np.append(data_ps, ps, axis=1)
    proxy_input['headers'] = [
        'period', 'region', 'table', 'origin', 'destination', 'value',
        'uncert', 'constraint', 'quantity', 'unit', 'factor', 'source', 'product', 'sector'
    ]
    return proxy_input


def write_proxy_output_in_excel(
    input_file: str,
    headers: list,
    sheet_name: str,
    proxy_output  # array with proxy output results
):
    try:
        act_xl = load_workbook(input_file)
        with pd.ExcelWriter(input_file, engine='openpyxl') as writer:
            df_data = pd.DataFrame(proxy_output, columns=headers[:12])
            writer.book = act_xl
            df_data.to_excel(writer, sheet_name=sheet_name, index=False)
        return True
    except Exception as expt:
        su_trace.logger.info('Exception ! Message : ' + str(expt))
        return False
